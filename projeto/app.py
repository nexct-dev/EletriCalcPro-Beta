import streamlit as st
import numpy as np
import pandas as pd
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from matplotlib.patches import FancyBboxPatch, Circle, FancyArrowPatch
try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
try:
    import ezdxf
    EZDXF_AVAILABLE = True
except ImportError:
    EZDXF_AVAILABLE = False

# Configuração página
st.set_page_config(
    page_title="Software Projetos Elétricos",
    page_icon="⚡",
    layout="wide",
    initial_sidebar_state="expanded"
)

# CSS personalizado
st.markdown("""
<style>
    .stTabs [data-baseweb="tab-list"] button [data-testid="stMarkdownContainer"] p {
        font-size: 1.1rem;
        font-weight: 500;
    }
    .metric-box {
        background-color: #f0f2f6;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
</style>
""", unsafe_allow_html=True)

# ============ TABELAS E DADOS - NBR 5410 ============

# Tabela 33 - Métodos de Instalação e Referências
metodos_instalacao = {
    'cond_visivel': {'codigo': 'A1', 'descricao': 'Condutor visível em alvenaria'},
    'eletroduto_embutido': {'codigo': 'B1', 'descricao': 'Eletroduto embutido em alvenaria'},
    'eletroduto_superficie': {'codigo': 'B2', 'descricao': 'Eletroduto na superfície'},
    'eletrocalha': {'codigo': 'C', 'descricao': 'Eletrocalha'},
    'bandeja': {'codigo': 'D', 'descricao': 'Bandeja'},
    'enterrado': {'codigo': 'E', 'descricao': 'Enterrado'},
}

# Tabela 36 - Capacidade de Condução de Corrente (A) - Cobre - PVC - 30°C ar
# Formato: (bitola_mm2): {metodo_ref: ampacidade}
tabela_36_cobre = {
    1.5: {'A1': 17.5, 'B1': 13.5, 'B2': 15.5, 'C': 17.5, 'D': 18, 'E': 17},
    2.5: {'A1': 24, 'B1': 18.5, 'B2': 21, 'C': 24, 'D': 25, 'E': 23},
    4: {'A1': 32, 'B1': 25, 'B2': 28, 'C': 32, 'D': 33, 'E': 31},
    6: {'A1': 41, 'B1': 32, 'B2': 36, 'C': 41, 'D': 43, 'E': 39},
    10: {'A1': 57, 'B1': 44, 'B2': 50, 'C': 57, 'D': 60, 'E': 54},
    16: {'A1': 76, 'B1': 59, 'B2': 68, 'C': 76, 'D': 80, 'E': 72},
    25: {'A1': 99, 'B1': 77, 'B2': 89, 'C': 99, 'D': 104, 'E': 94},
    35: {'A1': 125, 'B1': 97, 'B2': 112, 'C': 125, 'D': 132, 'E': 119},
    50: {'A1': 155, 'B1': 120, 'B2': 138, 'C': 155, 'D': 164, 'E': 148},
    70: {'A1': 194, 'B1': 150, 'B2': 172, 'C': 194, 'D': 206, 'E': 186},
    95: {'A1': 232, 'B1': 179, 'B2': 206, 'C': 232, 'D': 247, 'E': 223},
    120: {'A1': 263, 'B1': 203, 'B2': 234, 'C': 263, 'D': 280, 'E': 253},
    150: {'A1': 295, 'B1': 227, 'B2': 262, 'C': 295, 'D': 314, 'E': 284},
    185: {'A1': 327, 'B1': 252, 'B2': 291, 'C': 327, 'D': 349, 'E': 315},
    240: {'A1': 369, 'B1': 285, 'B2': 328, 'C': 369, 'D': 394, 'E': 356},
}

# Tabela 36 - Capacidade de Condução de Corrente (A) - Alumínio - PVC - 30°C ar
tabela_36_aluminio = {
    2.5: {'A1': 18, 'B1': 14, 'B2': 16, 'C': 18, 'D': 19, 'E': 17},
    4: {'A1': 24, 'B1': 19, 'B2': 21, 'C': 24, 'D': 25, 'E': 23},
    6: {'A1': 31, 'B1': 24, 'B2': 27, 'C': 31, 'D': 32, 'E': 29},
    10: {'A1': 43, 'B1': 33, 'B2': 38, 'C': 43, 'D': 45, 'E': 41},
    16: {'A1': 57, 'B1': 44, 'B2': 51, 'C': 57, 'D': 60, 'E': 54},
    25: {'A1': 75, 'B1': 58, 'B2': 67, 'C': 75, 'D': 79, 'E': 71},
    35: {'A1': 94, 'B1': 73, 'B2': 85, 'C': 94, 'D': 99, 'E': 90},
    50: {'A1': 117, 'B1': 90, 'B2': 104, 'C': 117, 'D': 124, 'E': 112},
    70: {'A1': 147, 'B1': 113, 'B2': 130, 'C': 147, 'D': 156, 'E': 141},
    95: {'A1': 176, 'B1': 135, 'B2': 156, 'C': 176, 'D': 187, 'E': 169},
    120: {'A1': 199, 'B1': 153, 'B2': 177, 'C': 199, 'D': 212, 'E': 191},
    150: {'A1': 224, 'B1': 172, 'B2': 198, 'C': 224, 'D': 238, 'E': 215},
    185: {'A1': 248, 'B1': 191, 'B2': 220, 'C': 248, 'D': 264, 'E': 239},
    240: {'A1': 280, 'B1': 215, 'B2': 248, 'C': 280, 'D': 298, 'E': 270},
}

# Tabela 42 - Fatores de Correção por Agrupamento de Condutores
# Número de circuitos no mesmo eletroduto/canaleta
fatores_agrupamento = {
    1: 1.0,
    2: 0.80,
    3: 0.70,
    4: 0.65,
    5: 0.60,
    6: 0.57,
    7: 0.54,
    8: 0.52,
    9: 0.50,
}

# Tabela 46 - Número de Condutores Carregados
condutores_carregados = {
    'monofasico_2f': 2,
    'monofasico_3f': 3,
    'trifasico_sem_neutro': 3,
    'trifasico_com_neutro': 4,
}

# Seções mínimas conforme uso
secoes_minimas = {
    'Iluminacao': 1.5,
    'Tomada_residencial': 2.5,
    'Tomada_forcados': 6,
    'Corrente_continua': 6
}

# Tamanhos padrão de transformadores
tamanhos_padrao_kva = [10, 15, 25, 30, 45, 50, 75, 100, 150, 200, 300, 500, 750, 1000]

# ============ TABELAS E DADOS - NBR 5419 (PROTEÇÃO CONTRA DESCARGAS ATMOSFÉRICAS) ============

# Tabela 1 - Relação entre Níveis de Proteção e Classes de SPDA (NBR 5419-3:2015)
niveis_protecao_spda = {
    'I': {'nivel': 'I', 'classe': 'I', 'eficiencia_min': 0.98},
    'II': {'nivel': 'II', 'classe': 'II', 'eficiencia_min': 0.95},
    'III': {'nivel': 'III', 'classe': 'III', 'eficiencia_min': 0.90},
    'IV': {'nivel': 'IV', 'classe': 'IV', 'eficiencia_min': 0.80},
}

# Tabela 2 - Valores de Raio da Esfera Rolante, Tamanho da Malha e Ângulo de Proteção
parametros_spda = {
    'I': {
        'raio_esfera_rolante': 20,      # metros
        'tamanho_malha': (5, 5),         # metros (5x5)
        'angulo_protecao': None,         # não aplicável
        'distancia_condutores': 10,      # metros
        'distancia_aneis': 10            # metros
    },
    'II': {
        'raio_esfera_rolante': 30,
        'tamanho_malha': (10, 10),
        'angulo_protecao': None,
        'distancia_condutores': 15,
        'distancia_aneis': 15
    },
    'III': {
        'raio_esfera_rolante': 45,
        'tamanho_malha': (15, 15),
        'angulo_protecao': None,
        'distancia_condutores': 20,
        'distancia_aneis': 20
    },
    'IV': {
        'raio_esfera_rolante': 60,
        'tamanho_malha': (20, 20),
        'angulo_protecao': None,
        'distancia_condutores': 25,
        'distancia_aneis': 25
    }
}

# Tabela 3 - Espessura Mínima de Chapas ou Tubulações Metálicas para SPDA (mm)
espessura_minima_materiais = {
    'cobre': {
        'espessura': 2.0,
        'condutor_minimo': 50,           # mm² seção mínima
        'densidade': 8.9,                # g/cm³
        'resistividade': 1.68e-8,        # Ω·m
    },
    'aluminio': {
        'espessura': 2.5,
        'condutor_minimo': 70,
        'densidade': 2.7,
        'resistividade': 2.65e-8,
    },
    'aco_galvanizado': {
        'espessura': 4.0,
        'condutor_minimo': 95,
        'densidade': 7.85,
        'resistividade': 1.1e-7,
    },
    'aco_inoxidavel': {
        'espessura': 2.0,
        'condutor_minimo': 50,
        'densidade': 7.5,
        'resistividade': 7.2e-7,
    }
}

# Tabela 5 - Materiais para SPDA e Condições de Utilização
materiais_spda = {
    'cobre': {
        'aplicacao': 'Geral',
        'vantagens': ['Alta condutividade', 'Durável', 'Não oxida facilmente'],
        'desvantagens': ['Custo elevado', 'Alvo de furtos'],
        'ambientes': ['Residencial', 'Comercial', 'Industrial', 'Hospitalar'],
    },
    'aluminio': {
        'aplicacao': 'Geral',
        'vantagens': ['Leve', 'Menos custoso que cobre', 'Fácil instalação'],
        'desvantagens': ['Menos resistente à corrosão', 'Menor condutividade'],
        'ambientes': ['Residencial', 'Comercial', 'Estruturas leves'],
    },
    'aco_galvanizado': {
        'aplicacao': 'Estruturas Industriais',
        'vantagens': ['Baixo custo', 'Resistente à corrosão', 'Estrutura robusta'],
        'desvantagens': ['Menor condutividade', 'Peso elevado'],
        'ambientes': ['Industrial', 'Galpões', 'Estruturas expostas'],
    },
    'aco_inoxidavel': {
        'aplicacao': 'Ambientes Agressivos',
        'vantagens': ['Alta durabilidade', 'Resistência química', 'Melhor aparência'],
        'desvantagens': ['Custo muito elevado', 'Complexo de instalar'],
        'ambientes': ['Marítimo', 'Químico', 'Alimentício', 'Hospitalar'],
    }
}

# ============ FUNÇÕES DE CÁLCULO ============

def obter_ampacidade(secao, material='cobre', metodo='A1'):
    """
    Obtém a ampacidade da Tabela 36 conforme material, seção e método de instalação.
    """
    tabela = tabela_36_cobre if material == 'cobre' else tabela_36_aluminio
    
    if secao not in tabela:
        return None
    
    return tabela[secao].get(metodo, None)


def obter_fator_agrupamento(num_circuitos):
    """
    Obtém o fator de correção para agrupamento conforme Tabela 42.
    """
    if num_circuitos <= 0:
        return 1.0
    
    return fatores_agrupamento.get(num_circuitos, fatores_agrupamento.get(9, 0.50))


def calcular_corrente_ajustada(corrente_calculada, num_circuitos=1, fator_temperatura=1.0, fator_agrupamento_manual=1.0):
    """
    Calcula a corrente ajustada considerando fatores de correção.
    Iz' = I / (fator_temperatura * fator_agrupamento)
    """
    fator_agrp = obter_fator_agrupamento(num_circuitos)
    fator_correcao_total = fator_temperatura * fator_agrp * fator_agrupamento_manual
    
    return corrente_calculada / fator_correcao_total


def obter_secoes_disponiveis(material='cobre'):
    """
    Retorna as seções disponíveis conforme o material.
    """
    tabela = tabela_36_cobre if material == 'cobre' else tabela_36_aluminio
    return sorted(list(tabela.keys()))


def dimensionar_transformador(potencia_total_kw, tensao_primaria=13800, tensao_secundaria=380, 
                              fator_demanda=0.8, margem_crescimento=0.2):
    """
    Dimensiona transformador conforme NBR 5356.
    """
    alertas = []
    
    potencia_demanda = potencia_total_kw * fator_demanda
    potencia_projeto = potencia_demanda * (1 + margem_crescimento)
    potencia_kva = potencia_projeto / 0.92  # Considerando FP=0.92
    
    # Encontrar tamanho padrão
    kva_selecionado = next((kva for kva in tamanhos_padrao_kva if kva >= potencia_kva), None)
    
    if kva_selecionado is None:
        alertas.append("Potência fora da faixa padrão. Consulte fabricante.")
        kva_selecionado = potencia_kva
    
    corrente_primaria = (kva_selecionado * 1000) / (tensao_primaria * np.sqrt(3))
    corrente_secundaria = (kva_selecionado * 1000) / (tensao_secundaria * np.sqrt(3))
    
    return {
        "potencia_demanda": round(potencia_demanda, 2),
        "potencia_projeto": round(potencia_projeto, 2),
        "kva_selecionado": kva_selecionado,
        "corrente_primaria": round(corrente_primaria, 2),
        "corrente_secundaria": round(corrente_secundaria, 2),
        "conforme": len(alertas) == 0,
        "alertas": alertas
    }


def dimensionar_condutor(corrente_circuito, comprimento_circuito=30, queda_tensao_max=3.0, 
                         tensao_nominal=380, tipo_instalacao='eletroduto_embutido', material='cobre',
                         num_circuitos=1, fator_temperatura=1.0):
    """
    Dimensiona condutor conforme NBR 5410 com tabelas reais.
    Critério de queda de tensão máxima + verificação de ampacidade.
    """
    alertas = []
    
    # Mapa de tipos de instalação para código NBR
    mapa_tipos = {
        'cond_visivel': 'A1',
        'eletroduto_embutido': 'B1',
        'eletroduto_superficie': 'B2',
        'eletrocalha': 'C',
        'bandeja': 'D',
        'enterrado': 'E',
    }
    
    codigo_metodo = mapa_tipos.get(tipo_instalacao, 'B1')
    
    # Resistividade (Ω·mm²/m) a 20°C
    rho = 0.0175 if material == 'cobre' else 0.029
    
    # Comprimento reduzido (ida e volta)
    comprimento_reduzido = 2 * comprimento_circuito
    
    # Queda de tensão máxima em V
    delta_u_max = (queda_tensao_max / 100) * tensao_nominal
    
    # Seção mínima pelo critério de queda de tensão
    # S = (rho * L_r * I) / delta_U_max
    secao_minima_queda = (rho * comprimento_reduzido * corrente_circuito) / delta_u_max
    
    # Corrente ajustada para fatores de correção
    corrente_ajustada = calcular_corrente_ajustada(
        corrente_circuito, 
        num_circuitos=num_circuitos,
        fator_temperatura=fator_temperatura
    )
    
    # Seções disponíveis
    secoes_disp = obter_secoes_disponiveis(material)
    
    # Encontrar seção que atende queda de tensão
    secao_por_queda = next((s for s in secoes_disp if s >= secao_minima_queda), None)
    
    if secao_por_queda is None:
        alertas.append(f"Seção calculada ({secao_minima_queda:.2f} mm²) acima do padrão.")
        secao_por_queda = secoes_disp[-1]
    
    # Encontrar seção que atende ampacidade
    secao_selecionada = None
    ampacidade_real = 0
    
    for secao in secoes_disp:
        if secao >= secao_por_queda:
            amp = obter_ampacidade(secao, material, codigo_metodo)
            if amp and amp >= corrente_ajustada:
                secao_selecionada = secao
                ampacidade_real = amp
                break
    
    if secao_selecionada is None:
        alertas.append(f"Corrente ajustada {corrente_ajustada:.2f}A acima da ampacidade máxima.")
        secao_selecionada = secoes_disp[-1]
        ampacidade_real = obter_ampacidade(secao_selecionada, material, codigo_metodo)
    
    # Verificar se queda real está dentro dos limites
    r_real = (rho * comprimento_reduzido) / secao_selecionada
    queda_real = (r_real * corrente_circuito * 100) / tensao_nominal
    
    if queda_real > queda_tensao_max:
        alertas.append(f"Queda real {queda_real:.2f}% > máximo {queda_tensao_max}%. Aumentar seção.")
    
    # Verificar se ampacidade sem fatores de correção atende
    if corrente_circuito > ampacidade_real:
        alertas.append(f"Corrente {corrente_circuito}A > Ampacidade {ampacidade_real}A (considerando fatores).")
    
    return {
        "secao_minima_calculada": round(secao_minima_queda, 2),
        "secao_selecionada": secao_selecionada,
        "ampacidade": ampacidade_real,
        "corrente_ajustada": round(corrente_ajustada, 2),
        "queda_tensao_real": round(queda_real, 2),
        "conforme": len(alertas) == 0,
        "alertas": alertas,
        "metodo_instalacao": codigo_metodo,
        "material": material.capitalize()
    }


def dimensionar_disjuntor(corrente_circuito, tipo_circuito='geral', padrao='c'):
    """
    Seleciona disjuntor conforme corrente do circuito e padrão.
    """
    alertas = []
    
    # Tamanhos padrão (A)
    tamanhos_c = [2, 6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250, 400]
    tamanhos_b = [2, 6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200]
    tamanhos_d = [2, 6, 10, 16, 20, 25, 32, 40, 50, 63, 80, 100, 125, 160, 200, 250, 400]
    
    tamanhos = tamanhos_c if padrao == 'c' else (tamanhos_b if padrao == 'b' else tamanhos_d)
    
    # Seleção: primeiro que cumpre Iz >= In
    corrente_nominal = next((t for t in tamanhos if t >= corrente_circuito), None)
    
    if corrente_nominal is None:
        alertas.append(f"Corrente {corrente_circuito}A fora da faixa disponível.")
        corrente_nominal = tamanhos[-1]
    
    # Verificação de proteção (In <= Iz)
    if corrente_nominal > corrente_circuito * 1.25:
        alertas.append(f"In {corrente_nominal}A > 1.25*Iz {corrente_circuito*1.25:.1f}A. Verificar seleção.")
    
    return {
        "corrente_nominal": corrente_nominal,
        "padrao": padrao.upper(),
        "tipo": tipo_circuito,
        "conforme": len(alertas) == 0,
        "alertas": alertas
    }


def calcular_curto_circuito(kva_transformador, tensao_secundaria=380, uk_percent=5.0, comprimento_cabo=0, 
                           secao_cabo=0, rho_cabo=0.023, x_cabo_unit=0.00008, tipo_curto='trifasico'):
    """
    Calcula corrente de curto-circuito conforme IEC 60909/NBR 5410.
    """
    import unicodedata
    alertas = []
    
    # Normalizar tipo_curto removendo acentos
    tipo_curto_norm = unicodedata.normalize('NFKD', tipo_curto.lower()).encode('ASCII', 'ignore').decode('ASCII')
    
    uk = uk_percent / 100
    zk_trafo = (tensao_secundaria ** 2) / (kva_transformador * 1000) * uk
    
    if tipo_curto_norm == 'trifasico':
        ik_sec = (tensao_secundaria / np.sqrt(3)) / zk_trafo / 1000
        fator = 1.0
    elif tipo_curto_norm == 'bifasico':
        ik_sec = (tensao_secundaria / 2) / zk_trafo / 1000
        fator = np.sqrt(3)/2
    elif tipo_curto_norm == 'monofasico':
        ik_sec = tensao_secundaria / zk_trafo / 1000
        fator = 1.5
    else:
        alertas.append("Tipo de curto inválido.")
        return {"alertas": alertas}
    
    ik_ponto = ik_sec
    
    if comprimento_cabo > 0 and secao_cabo > 0:
        r_cabo = (rho_cabo * comprimento_cabo) / secao_cabo
        x_cabo = x_cabo_unit * comprimento_cabo
        z_cabo = np.sqrt(r_cabo**2 + x_cabo**2)
        z_total = np.sqrt(zk_trafo**2 + z_cabo**2)
        ik_ponto = (tensao_secundaria / np.sqrt(3)) / z_total / 1000 * fator
    
    conforme = len(alertas) == 0
    return {
        "ik_secundario": ik_sec,
        "ik_ponto": ik_ponto,
        "conforme": conforme,
        "alertas": alertas
    }


def balancear_fases(cargas_fase_a, cargas_fase_b, cargas_fase_c, tensao=380):
    """
    Realiza balanceamento de fases trifásico.
    Entrada: Lista de cargas (potências em kW) para cada fase.
    """
    alertas = []
    
    # Somar cargas por fase
    soma_a = sum(cargas_fase_a)
    soma_b = sum(cargas_fase_b)
    soma_c = sum(cargas_fase_c)
    
    total_cargas = soma_a + soma_b + soma_c
    
    if total_cargas == 0:
        alertas.append("Nenhuma carga definida.")
        return {
            "cargas_a": 0, "cargas_b": 0, "cargas_c": 0,
            "correntes_a": 0, "correntes_b": 0, "correntes_c": 0,
            "desbalanceamento": 0, "conforme": False, "alertas": alertas
        }
    
    # Cálculo de correntes (I = P / (√3 × V × FP), considerando FP=0.92)
    fp = 0.92
    i_a = (soma_a * 1000) / (np.sqrt(3) * tensao * fp)
    i_b = (soma_b * 1000) / (np.sqrt(3) * tensao * fp)
    i_c = (soma_c * 1000) / (np.sqrt(3) * tensao * fp)
    
    # Corrente média
    i_media = (i_a + i_b + i_c) / 3
    
    # Desbalanceamento em percentual
    # NBR 5410: máximo 3% para baixa tensão
    if i_media > 0:
        desbalanc = ((max(i_a, i_b, i_c) - min(i_a, i_b, i_c)) / i_media) * 100
    else:
        desbalanc = 0
    
    # Verificação de conformidade
    conforme = desbalanc <= 3.0
    
    if not conforme:
        alertas.append(f"Desbalanceamento {desbalanc:.1f}% > 3% (máx NBR 5410).")
        alertas.append("Recomenda-se redistribuir cargas entre fases.")
    
    # Sugestão de balanceamento
    carga_media = total_cargas / 3
    desvio_a = soma_a - carga_media
    desvio_b = soma_b - carga_media
    desvio_c = soma_c - carga_media
    
    return {
        "cargas_a": round(soma_a, 2),
        "cargas_b": round(soma_b, 2),
        "cargas_c": round(soma_c, 2),
        "correntes_a": round(i_a, 2),
        "correntes_b": round(i_b, 2),
        "correntes_c": round(i_c, 2),
        "corrente_media": round(i_media, 2),
        "desbalanceamento": round(desbalanc, 2),
        "carga_media": round(carga_media, 2),
        "desvio_a": round(desvio_a, 2),
        "desvio_b": round(desvio_b, 2),
        "desvio_c": round(desvio_c, 2),
        "conforme": conforme,
        "alertas": alertas
    }


def gerar_unifilar_matplotlib(resultado_condutor, nome_circuito="Circuito"):
    """
    Gera diagrama unifilar em matplotlib.
    """
    fig, ax = plt.subplots(figsize=(14, 8))
    ax.set_xlim(0, 14)
    ax.set_ylim(0, 8)
    ax.axis('off')
    
    # Cor de fundo
    fig.patch.set_facecolor('white')
    
    # Título
    ax.text(7, 7.5, f"Diagrama Unifilar - {nome_circuito}", 
            fontsize=16, fontweight='bold', ha='center')
    
    # Fonte de alimentação (Trafo)
    trafo_box = FancyBboxPatch((0.5, 5.5), 1.5, 1, 
                               boxstyle="round,pad=0.1", 
                               edgecolor='black', facecolor='lightblue', linewidth=2)
    ax.add_patch(trafo_box)
    ax.text(1.25, 6, "TRAFO\n380V", ha='center', va='center', fontsize=9, fontweight='bold')
    
    # Barramento principal
    ax.plot([2.2, 12], [6, 6], 'k-', linewidth=3)
    ax.text(1.2, 6.3, "Barramento Principal", fontsize=9, style='italic')
    
    # Fases A, B, C (cores padrão)
    cores_fases = {'A': 'red', 'B': 'yellow', 'C': 'blue'}
    posicoes_y = [6.5, 6, 5.5]
    fases = ['A', 'B', 'C']
    
    for idx, (fase, y) in enumerate(zip(fases, posicoes_y)):
        # Identificação da fase
        circle = Circle((0.5, y), 0.15, color=cores_fases[fase], zorder=2)
        ax.add_patch(circle)
        ax.text(0.3, y, fase, ha='center', va='center', fontsize=8, fontweight='bold', color='white')
        
        # Disjuntor
        disj_x = 3 + idx * 3
        ax.plot([2.2, disj_x-0.3], [y, y], color=cores_fases[fase], linewidth=2)
        ax.plot([disj_x-0.3, disj_x+0.3], [y-0.15, y+0.15], 'k-', linewidth=2)
        ax.text(disj_x, y+0.35, f"Disj.\n{resultado_condutor['ampacidade']}A", 
                ha='center', va='bottom', fontsize=8)
        
        # Condutor até carga
        ax.plot([disj_x+0.3, disj_x+2], [y, y], color=cores_fases[fase], linewidth=2)
        
        # Caixa de carga
        carga_box = FancyBboxPatch((disj_x+2, y-0.25), 1.2, 0.5,
                                   boxstyle="round,pad=0.05",
                                   edgecolor=cores_fases[fase], 
                                   facecolor='lightyellow', linewidth=1.5)
        ax.add_patch(carga_box)
        ax.text(disj_x+2.6, y, "CARGA", ha='center', va='center', fontsize=7, fontweight='bold')
    
    # Informações do condutor
    info_text = f"""
Condutor: {resultado_condutor['secao_selecionada']} mm² ({resultado_condutor['material']})
Ampacidade: {resultado_condutor['ampacidade']} A
Queda Tensão: {resultado_condutor['queda_tensao_real']:.2f}%
Método Instalação: {resultado_condutor['metodo_instalacao']}
    """
    ax.text(10.5, 3.5, info_text.strip(), fontsize=9, 
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.8),
            verticalalignment='top', family='monospace')
    
    # Aterramento
    ax.plot([7, 7], [5, 4.5], 'g-', linewidth=2)
    for i in range(4):
        ax.plot([6.7, 7.3], [4.5 - i*0.15, 4.5 - i*0.15], 'g-', linewidth=2)
    ax.text(7.5, 4.2, "GND", fontsize=8, color='green', fontweight='bold')
    
    # Legenda
    legenda_text = f"Data: {datetime.now().strftime('%d/%m/%Y')}\nNBR 5410"
    ax.text(0.5, 0.5, legenda_text, fontsize=8, style='italic')
    
    return fig


def gerar_pdf_unifilar(resultado_condutor, nome_circuito="Circuito"):
    """
    Gera PDF com diagrama unifilar usando reportlab.
    """
    if not REPORTLAB_AVAILABLE:
        return None
    
    output = BytesIO()
    pdf_canvas = canvas.Canvas(output, pagesize=A4)
    
    # Dimensões A4
    width, height = A4
    
    # Cabeçalho
    pdf_canvas.setFont("Helvetica-Bold", 16)
    pdf_canvas.drawString(50, height-50, f"Diagrama Unifilar - {nome_circuito}")
    
    # Linha horizontal
    pdf_canvas.setLineWidth(2)
    pdf_canvas.line(50, height-100, width-50, height-100)
    
    # Informações do projeto
    pdf_canvas.setFont("Helvetica", 10)
    y_pos = height - 150
    info_lines = [
        f"Seção do Condutor: {resultado_condutor['secao_selecionada']} mm² ({resultado_condutor['material']})",
        f"Ampacidade: {resultado_condutor['ampacidade']} A",
        f"Corrente Ajustada: {resultado_condutor['corrente_ajustada']:.2f} A",
        f"Queda de Tensão: {resultado_condutor['queda_tensao_real']:.2f}%",
        f"Método de Instalação: {resultado_condutor['metodo_instalacao']}",
        f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}",
        f"Norma: NBR 5410",
    ]
    
    for line in info_lines:
        pdf_canvas.drawString(50, y_pos, line)
        y_pos -= 20
    
    # Diagrama simples de fases
    y_diag = y_pos - 40
    fases_info = [
        ("FASE A (Vermelho)", 50),
        ("FASE B (Amarelo)", 150),
        ("FASE C (Azul)", 250),
    ]
    
    pdf_canvas.setFont("Helvetica-Bold", 11)
    for fase_text, x_pos in fases_info:
        pdf_canvas.setFillColorRGB(0.8, 0.8, 0.8)
        pdf_canvas.rect(x_pos, y_diag-30, 80, 30, fill=True)
        pdf_canvas.setFillColorRGB(0, 0, 0)
        pdf_canvas.drawString(x_pos+5, y_diag-15, fase_text)
    
    # Rodapé
    pdf_canvas.setFont("Helvetica", 8)
    pdf_canvas.drawString(50, 30, f"© EletriCalcPro - Projetos Elétricos | {datetime.now().strftime('%d/%m/%Y')}")
    
    pdf_canvas.save()
    output.seek(0)
    return output


def gerar_dwg_unifilar(resultado_condutor, nome_circuito="Circuito"):
    """
    Gera DWG com diagrama unifilar usando ezdxf.
    """
    if not EZDXF_AVAILABLE:
        return None
    
    # Criar novo documento DWG
    doc = ezdxf.new('R2010')
    msp = doc.modelspace()
    
    # Camadas
    doc.layers.new(name='Fases', dxfattribs={'color': 7})
    doc.layers.new(name='Disjuntores', dxfattribs={'color': 3})
    doc.layers.new(name='Condutores', dxfattribs={'color': 1})
    doc.layers.new(name='Texto', dxfattribs={'color': 256})
    
    # Título
    msp.add_text(f"DIAGRAMA UNIFILAR - {nome_circuito}", dxfattribs={
        'height': 5, 'layer': 'Texto'
    }).set_pos((0, 100))
    
    # Trafo
    msp.add_lwpolyline([
        (0, 50), (10, 50), (10, 80), (0, 80), (0, 50)
    ], dxfattribs={'layer': 'Fases'})
    msp.add_text("TRAFO\n380V", dxfattribs={'height': 2, 'layer': 'Texto'}).set_pos((2, 60))
    
    # Barramento principal
    msp.add_line((15, 65), (150, 65), dxfattribs={'layer': 'Condutores', 'lineweight': 30})
    
    # Fases
    fases = ['A', 'B', 'C']
    cores_ezdxf = [1, 2, 5]  # Red, Yellow, Blue em DWG
    
    for idx, (fase, cor) in enumerate(zip(fases, cores_ezdxf)):
        y_fase = 80 - idx * 15
        
        # Ponto de fase
        msp.add_circle((5, y_fase), 2, dxfattribs={'color': cor, 'layer': 'Fases'})
        msp.add_text(fase, dxfattribs={'height': 1.5, 'layer': 'Texto', 'color': 7}).set_pos((3.5, y_fase-0.5))
        
        # Disjuntor
        x_disj = 35 + idx * 40
        msp.add_lwpolyline([
            (x_disj-3, y_fase-2), (x_disj+3, y_fase+2)
        ], dxfattribs={'layer': 'Disjuntores', 'lineweight': 15})
        msp.add_text(f"Disj\n{resultado_condutor['ampacidade']}A", 
                     dxfattribs={'height': 1.5, 'layer': 'Texto'}).set_pos((x_disj-2, y_fase+3))
        
        # Condutor até carga
        msp.add_line((x_disj+5, y_fase), (x_disj+25, y_fase), 
                     dxfattribs={'layer': 'Condutores', 'color': cor, 'lineweight': 20})
        
        # Caixa de carga
        msp.add_lwpolyline([
            (x_disj+25, y_fase-3), (x_disj+35, y_fase-3),
            (x_disj+35, y_fase+3), (x_disj+25, y_fase+3),
            (x_disj+25, y_fase-3)
        ], dxfattribs={'layer': 'Fases', 'color': cor})
        msp.add_text("CARGA", dxfattribs={'height': 1, 'layer': 'Texto'}).set_pos((x_disj+30, y_fase))
    
    # Informações do condutor
    info_y = 30
    info_lines = [
        f"Seção: {resultado_condutor['secao_selecionada']} mm² ({resultado_condutor['material']})",
        f"Ampacidade: {resultado_condutor['ampacidade']} A",
        f"Corrente Ajustada: {resultado_condutor['corrente_ajustada']:.2f} A",
        f"Queda Tensão: {resultado_condutor['queda_tensao_real']:.2f}%",
        f"Método: {resultado_condutor['metodo_instalacao']}",
    ]
    
    for idx, line in enumerate(info_lines):
        msp.add_text(line, dxfattribs={'height': 2, 'layer': 'Texto'}).set_pos((15, info_y - idx*5))
    
    # Aterramento
    msp.add_line((75, 35), (75, 20), dxfattribs={'layer': 'Fases', 'color': 3})
    for i in range(3):
        msp.add_line((70, 20-i*3), (80, 20-i*3), dxfattribs={'layer': 'Fases', 'color': 3})
    
    # Rodapé
    msp.add_text(f"© EletriCalcPro - NBR 5410 | {datetime.now().strftime('%d/%m/%Y')}", 
                 dxfattribs={'height': 1.5, 'layer': 'Texto'}).set_pos((0, 0))
    
    # Salvar em BytesIO
    output = BytesIO()
    doc.write(output)
    output.seek(0)
    return output


def gerar_relatorio(resultado, tipo='condutor', **kwargs):
    """
    Gera memorial descritivo conforme normas com referência às tabelas NBR 5410.
    """
    rel = f"{'='*60}\n"
    rel += f"MEMORIAL DESCRITIVO - {tipo.upper()}\n"
    rel += f"Normas: NBR 5410 / NBR 5356 / IEC 60909\n"
    rel += f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
    rel += f"{'='*60}\n\n"
    
    if tipo == 'condutor':
        rel += "PARÂMETROS DE ENTRADA\n"
        rel += f"{'─'*60}\n"
        rel += f"Corrente do Circuito: {kwargs.get('corrente_circuito', 0):.2f} A\n"
        rel += f"Comprimento: {kwargs.get('comprimento_circuito', 0):.1f} m\n"
        rel += f"Queda de Tensão Máxima: {kwargs.get('queda_tensao_max', 3)}%\n"
        rel += f"Tensão Nominal: {kwargs.get('tensao_nominal', 380):.0f} V\n"
        rel += f"Material: {resultado.get('material', 'Cobre')}\n"
        rel += f"Tipo de Instalação (Tabela 33): {resultado.get('metodo_instalacao', 'B1')}\n"
        rel += f"Nº de Circuitos Agrupados (Tabela 42): {kwargs.get('num_circuitos', 1)}\n\n"
        
        rel += "RESULTADOS DO CÁLCULO\n"
        rel += f"{'─'*60}\n"
        rel += f"Seção Mínima Calculada: {resultado['secao_minima_calculada']:.2f} mm²\n"
        rel += f"Seção Selecionada: {resultado['secao_selecionada']} mm²\n"
        rel += f"Ampacidade (Tabela 36): {resultado['ampacidade']} A\n"
        rel += f"Corrente Ajustada: {resultado.get('corrente_ajustada', resultado['ampacidade']):.2f} A\n"
        rel += f"Queda de Tensão Real: {resultado['queda_tensao_real']:.2f}%\n\n"
        
        rel += "REFERÊNCIAS NORMATIVAS\n"
        rel += f"{'─'*60}\n"
        rel += "• Tabela 33: Métodos de instalação conforme NBR 5410\n"
        rel += "• Tabela 36: Capacidade de condução de corrente\n"
        rel += "• Tabela 42: Fatores de correção por agrupamento\n"
        rel += "• Tabela 46: Número de condutores carregados\n"
        
    elif tipo == 'transformador':
        rel += f"Potência Total: {kwargs.get('potencia_total_kw', 0):.2f} kW\n"
        rel += f"Fator de Demanda: {kwargs.get('fator_demanda', 0.8):.1%}\n"
        rel += f"Margem de Crescimento: {kwargs.get('margem_crescimento', 0.2):.1%}\n\n"
        rel += f"Potência Demanda: {resultado['potencia_demanda']:.2f} kW\n"
        rel += f"Potência Projeto: {resultado['potencia_projeto']:.2f} kW\n"
        rel += f"Transformador Selecionado: {resultado['kva_selecionado']} kVA\n"
        rel += f"Corrente Primária: {resultado['corrente_primaria']:.2f} A\n"
        rel += f"Corrente Secundária: {resultado['corrente_secundaria']:.2f} A\n"
        
    elif tipo == 'disjuntor':
        rel += f"Corrente do Circuito: {kwargs.get('corrente_circuito', 0):.2f} A\n"
        rel += f"Tipo de Circuito: {kwargs.get('tipo_circuito', 'geral')}\n\n"
        rel += f"Padrão Selecionado: {resultado['padrao']}\n"
        rel += f"Corrente Nominal: {resultado['corrente_nominal']} A\n"
        
    elif tipo == 'curto_circuito':
        rel += f"Transformador: {kwargs.get('kva_transformador', 0)} kVA\n"
        rel += f"Tensão Secundária: {kwargs.get('tensao_secundaria', 380)} V\n"
        rel += f"Impedância Uk: {kwargs.get('uk_percent', 5)}%\n"
        rel += f"Tipo de Curto: {kwargs.get('tipo_curto', 'trifásico')}\n\n"
        rel += f"Ik no Secundário: {resultado['ik_secundario']:.2f} kA\n"
        rel += f"Ik no Ponto: {resultado['ik_ponto']:.2f} kA\n"
    
    if resultado.get('alertas', []):
        rel += "\nALERTAS:\n"
        for i, alerta in enumerate(resultado['alertas'], 1):
            rel += f"  {i}. {alerta}\n"
    else:
        rel += "\n✓ Cálculo conforme as normas aplicáveis.\n"
    
    rel += f"\n{'='*60}\n"
    return rel


def exportar_excel(resultado, tipo='condutor', **kwargs):
    """
    Exporta resultado para arquivo Excel com formatação profissional.
    """
    output = BytesIO()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = tipo.capitalize()
    
    # Estilos
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    title_font = Font(bold=True, size=14, color="366092")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Título
    ws.merge_cells('A1:D1')
    title = ws['A1']
    title.value = f"Cálculo de {tipo.capitalize()} - NBR 5410/5356"
    title.font = title_font
    
    # Data
    ws.merge_cells('A2:D2')
    date_cell = ws['A2']
    date_cell.value = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    date_cell.font = Font(italic=True, size=10)
    
    row = 4
    
    # ENTRADA
    ws.merge_cells(f'A{row}:D{row}')
    input_header = ws[f'A{row}']
    input_header.value = "PARÂMETROS DE ENTRADA"
    input_header.font = header_font
    input_header.fill = header_fill
    row += 1
    
    if tipo == 'condutor':
        dados_entrada = [
            ['Corrente do Circuito', f"{kwargs.get('corrente_circuito', 0):.2f}", "A"],
            ['Comprimento', f"{kwargs.get('comprimento_circuito', 0):.1f}", "m"],
            ['Queda Tensão Máx.', f"{kwargs.get('queda_tensao_max', 3)}", "%"],
            ['Tensão Nominal', f"{kwargs.get('tensao_nominal', 380):.0f}", "V"],
            ['Material', kwargs.get('material', 'Cobre'), ""],
            ['Tipo Instalação', kwargs.get('tipo_instalacao', 'Eletroduto embutido'), ""],
            ['Nº de Circuitos', f"{kwargs.get('num_circuitos', 1)}", ""],
        ]
    elif tipo == 'transformador':
        dados_entrada = [
            ['Potência Total', f"{kwargs.get('potencia_total_kw', 0):.2f}", "kW"],
            ['Fator Demanda', f"{kwargs.get('fator_demanda', 0.8):.1%}", "-"],
            ['Margem Crescimento', f"{kwargs.get('margem_crescimento', 0.2):.1%}", "-"],
            ['Tensão Primária', f"{kwargs.get('tensao_primaria', 13800):.0f}", "V"],
            ['Tensão Secundária', f"{kwargs.get('tensao_secundaria', 380):.0f}", "V"],
        ]
    elif tipo == 'disjuntor':
        dados_entrada = [
            ['Corrente Circuito', f"{kwargs.get('corrente_circuito', 0):.2f}", "A"],
            ['Tipo Circuito', kwargs.get('tipo_circuito', 'geral'), "-"],
        ]
    elif tipo == 'curto_circuito':
        dados_entrada = [
            ['Potência Trafo', f"{kwargs.get('kva_transformador', 0):.0f}", "kVA"],
            ['Tensão Secundária', f"{kwargs.get('tensao_secundaria', 380):.0f}", "V"],
            ['Impedância Uk', f"{kwargs.get('uk_percent', 5):.1f}", "%"],
            ['Tipo Curto', kwargs.get('tipo_curto', 'trifásico'), "-"],
        ]
    
    for param, valor, unidade in dados_entrada:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = valor
        ws[f'C{row}'] = unidade
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = border
            ws[f'{col}{row}'].alignment = Alignment(horizontal='left')
        row += 1
    
    row += 1
    
    # RESULTADOS
    ws.merge_cells(f'A{row}:D{row}')
    output_header = ws[f'A{row}']
    output_header.value = "RESULTADOS"
    output_header.font = header_font
    output_header.fill = header_fill
    row += 1
    
    if tipo == 'condutor':
        dados_saida = [
            ['Seção Mínima Calculada', f"{resultado['secao_minima_calculada']:.2f}", "mm²"],
            ['Seção Selecionada', f"{resultado['secao_selecionada']}", "mm²"],
            ['Ampacidade (Tabela 36)', f"{resultado['ampacidade']}", "A"],
            ['Corrente Ajustada', f"{resultado['corrente_ajustada']:.2f}", "A"],
            ['Queda Tensão Real', f"{resultado['queda_tensao_real']:.2f}", "%"],
            ['Método Instalação', resultado['metodo_instalacao'], ""],
            ['Material', resultado['material'], ""],
            ['Conforme NBR 5410', "SIM" if resultado['conforme'] else "NÃO", ""],
        ]
    elif tipo == 'transformador':
        dados_saida = [
            ['Potência Demanda', f"{resultado['potencia_demanda']:.2f}", "kW"],
            ['Potência Projeto', f"{resultado['potencia_projeto']:.2f}", "kW"],
            ['Transformador', f"{resultado['kva_selecionado']}", "kVA"],
            ['Corrente Primária', f"{resultado['corrente_primaria']:.2f}", "A"],
            ['Corrente Secundária', f"{resultado['corrente_secundaria']:.2f}", "A"],
            ['Conforme NBR 5356', "SIM" if resultado['conforme'] else "NÃO", ""],
        ]
    elif tipo == 'disjuntor':
        dados_saida = [
            ['Padrão', resultado['padrao'], ""],
            ['Corrente Nominal', f"{resultado['corrente_nominal']}", "A"],
            ['Tipo', resultado['tipo'].title(), ""],
            ['Conforme', "SIM" if resultado['conforme'] else "NÃO", ""],
        ]
    elif tipo == 'curto_circuito':
        dados_saida = [
            ['Ik Secundário', f"{resultado['ik_secundario']:.2f}", "kA"],
            ['Ik no Ponto', f"{resultado['ik_ponto']:.2f}", "kA"],
            ['Conforme IEC 60909', "SIM" if resultado['conforme'] else "NÃO", ""],
        ]
    
    for param, valor, unidade in dados_saida:
        ws[f'A{row}'] = param
        ws[f'B{row}'] = valor
        ws[f'C{row}'] = unidade
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.border = border
            cell.alignment = Alignment(horizontal='left')
            if 'Conforme' in param:
                cell.font = Font(bold=True)
        row += 1
    
    # Alertas
    if resultado.get('alertas', []):
        row += 1
        ws.merge_cells(f'A{row}:D{row}')
        alert_header = ws[f'A{row}']
        alert_header.value = "ALERTAS"
        alert_header.font = Font(color="FFFFFF", bold=True, size=11)
        alert_header.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
        row += 1
        
        for i, alerta in enumerate(resultado['alertas'], 1):
            ws.merge_cells(f'A{row}:D{row}')
            alert_cell = ws[f'A{row}']
            alert_cell.value = f"{i}. {alerta}"
            alert_cell.font = Font(color="C00000")
            alert_cell.border = border
            row += 1
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    
    wb.save(output)
    output.seek(0)
    return output


# ============ FUNÇÕES ESPECÍFICAS PARA NBR 5419 ============

def dimensionar_spda(classe_protecao, altura_estrutura, comprimento_estrutura, 
                     largura_estrutura, material_spda='cobre', tipo_metodo='esfera_rolante'):
    """
    Dimensiona sistema de proteção contra descargas atmosféricas conforme NBR 5419.
    
    Parâmetros:
    - classe_protecao: 'I', 'II', 'III' ou 'IV'
    - altura_estrutura: altura em metros
    - comprimento_estrutura: comprimento em metros
    - largura_estrutura: largura em metros
    - material_spda: 'cobre', 'aluminio', 'aco_galvanizado', 'aco_inoxidavel'
    - tipo_metodo: 'esfera_rolante', 'malha', 'angulo_protecao'
    """
    alertas = []
    
    # Verificar classe válida
    if classe_protecao not in parametros_spda:
        alertas.append(f"Classe {classe_protecao} não reconhecida. Use I, II, III ou IV.")
        classe_protecao = 'I'
    
    params = parametros_spda[classe_protecao]
    nivel_info = niveis_protecao_spda[classe_protecao]
    material_info = espessura_minima_materiais.get(material_spda, espessura_minima_materiais['cobre'])
    
    # Calcular número de condutores de descida necessários
    perimetro = 2 * (comprimento_estrutura + largura_estrutura)
    distancia_condutores = params['distancia_condutores']
    num_condutores_descida = max(2, int(np.ceil(perimetro / distancia_condutores)))
    
    # Calcular número de anéis condutores
    num_aneis = max(1, int(np.ceil(altura_estrutura / 20)))  # Um anel a cada 20m
    distancia_aneis = params['distancia_aneis']
    
    # Cálculo do raio da esfera rolante (Método 1: esfera rolante)
    raio_esfera = params['raio_esfera_rolante']
    
    # Verificar proteção baseado no método
    altura_captacao_recomendada = altura_estrutura + 2.0  # Captores 2m acima
    
    if tipo_metodo == 'esfera_rolante':
        # Verificar se a esfera rolante cobre toda a estrutura
        protecao_lateral = raio_esfera
        protecao_altura = raio_esfera
        
        if altura_estrutura > raio_esfera:
            alertas.append(f"Altura da estrutura ({altura_estrutura}m) > raio da esfera ({raio_esfera}m). "
                         f"Considere aumentar nível de proteção ou usar capta-raios adicionais.")
    
    elif tipo_metodo == 'malha':
        malha_x, malha_y = params['tamanho_malha']
        num_linhas_x = int(np.ceil(comprimento_estrutura / malha_x))
        num_linhas_y = int(np.ceil(largura_estrutura / malha_y))
        comprimento_malha_total = 2 * (num_linhas_x * largura_estrutura + num_linhas_y * comprimento_estrutura)
    else:
        # Método do ângulo de proteção (não aplicável para Classes I-IV no Brasil)
        alertas.append("Método de ângulo de proteção não aplicável para Classes I-IV.")
    
    # Cálculo de condutores necessários
    comprimento_condutor_descida = altura_estrutura * num_condutores_descida
    comprimento_condutor_malha = perimetro + (num_aneis * perimetro)
    comprimento_total_condutor = comprimento_condutor_descida + comprimento_condutor_malha
    
    # Cálculo da seção mínima de condutor (maior entre todas as limitações)
    secao_minima = material_info['condutor_minimo']
    
    # Resistência de aterramento recomendada (< 10 Ω)
    resistencia_aterramento_max = 10.0  # ohms
    
    # Calcular volume de cobre equivalente
    volume_material = (comprimento_total_condutor * secao_minima) / 1000  # m³ aprox
    massa_material = volume_material * material_info['densidade'] * 1e6  # kg
    
    return {
        'classe_protecao': classe_protecao,
        'nivel_protecao': nivel_info['nivel'],
        'eficiencia_minima': f"{nivel_info['eficiencia_min'] * 100:.0f}%",
        'raio_esfera_rolante': raio_esfera,
        'tamanho_malha': f"{params['tamanho_malha'][0]} × {params['tamanho_malha'][1]} m",
        'num_condutores_descida': num_condutores_descida,
        'distancia_entre_condutores': distancia_condutores,
        'num_aneis_condutores': num_aneis,
        'distancia_entre_aneis': distancia_aneis,
        'material_spda': material_spda.replace('_', ' ').title(),
        'espessura_minima': material_info['espessura'],
        'secao_minima_condutor': secao_minima,
        'comprimento_condutor_descida': round(comprimento_condutor_descida, 2),
        'comprimento_condutor_malha': round(comprimento_condutor_malha, 2),
        'comprimento_total_condutor': round(comprimento_total_condutor, 2),
        'massa_aproximada_material': round(massa_material, 2),
        'resistencia_aterramento_recomendada': f"< {resistencia_aterramento_max} Ω",
        'altura_captacao_recomendada': altura_captacao_recomendada,
        'conforme': len(alertas) == 0,
        'alertas': alertas
    }


def verificar_equipotencializacao(tensao_toque_limite=50, impedancia_corpo=1000):
    """
    Verifica requisitos de equipotencialização conforme NBR 5419-6:2015.
    
    Parâmetros:
    - tensao_toque_limite: Tensão segura de toque em V (padrão 50V para ambiente seco)
    - impedancia_corpo: Impedância do corpo humano em Ω (padrão 1000Ω)
    """
    
    # Corrente máxima segura
    corrente_segura = tensao_toque_limite / impedancia_corpo
    
    # Tempo de exposição seguro (Curva de Dalziel)
    tempo_seguro = 0.165 / np.sqrt(corrente_segura)  # segundos, para I < 100mA
    
    return {
        'tensao_toque_limite': tensao_toque_limite,
        'impedancia_corpo': impedancia_corpo,
        'corrente_segura': round(corrente_segura * 1000, 2),  # em mA
        'tempo_exposicao_seguro': round(tempo_seguro, 3),  # segundos
        'recomendacoes': [
            'Equipotencializar todos os condutores de descida',
            'Conectar equipamentos metálicos próximos à estrutura',
            'Usar DPS (Dispositivo de Proteção contra Surtos)',
            'Aterramento com resistência < 10 Ω'
        ]
    }


def calcular_corrente_descarga(energia_relativa, impedancia_arco=50):
    """
    Estima a corrente de descarga atmosférica para cálculos de proteção.
    
    Baseado em estatísticas de descargas em território brasileiro (35 mil/ano aprox).
    """
    
    # Valores típicos de corrente de descarga (kA)
    correntes_tipicas = {
        'minima': 5,
        'media': 25,
        'maxima': 200
    }
    
    # Corrente calculada baseada em energia relativa
    corrente_estimada = correntes_tipicas['media'] * (energia_relativa / 100)
    
    return {
        'corrente_minima_design': correntes_tipicas['minima'],
        'corrente_media_design': correntes_tipicas['media'],
        'corrente_maxima_design': correntes_tipicas['maxima'],
        'corrente_estimada': round(corrente_estimada, 2),
        'unidade': 'kA',
        'impedancia_arco_tipica': impedancia_arco,
    }


# ============ INICIALIZAÇÃO DE SESSÃO ============
# Inicializar projetos em session_state se não existir
if 'projetos' not in st.session_state:
    st.session_state.projetos = {}
if 'projeto_atual' not in st.session_state:
    st.session_state.projeto_atual = None
if 'historico_calculos' not in st.session_state:
    st.session_state.historico_calculos = []
if 'modo_sidebar' not in st.session_state:
    st.session_state.modo_sidebar = "Novo Projeto"

# ============ SIDEBAR - PROJETO & NAVEGAÇÃO ============
with st.sidebar:
    st.title("📁 Sistema de Projetos")
    st.divider()
    
    # Abas do sidebar
    sidebar_tab1, sidebar_tab2, sidebar_tab3 = st.tabs(["Novo", "Carregador", "Ferramentas"])
    
    with sidebar_tab1:
        st.subheader("📝 Novo Projeto")
        nome_novo_projeto = st.text_input("Nome do Projeto", placeholder="Ex: Edifício - Andar 5", key="sb_nome_projeto")
        cliente = st.text_input("Cliente", placeholder="Ex: Empresa XYZ", key="sb_cliente")
        local = st.text_input("Local", placeholder="Ex: São Paulo - SP", key="sb_local")
        
        if st.button("✅ Criar Projeto", use_container_width=True, key="btn_criar_proj"):
            if nome_novo_projeto.strip():
                novo_projeto = {
                    'nome': nome_novo_projeto,
                    'cliente': cliente,
                    'local': local,
                    'data_criacao': datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
                    'modulos': {}
                }
                st.session_state.projetos[nome_novo_projeto] = novo_projeto
                st.session_state.projeto_atual = nome_novo_projeto
                st.success(f"✅ Projeto '{nome_novo_projeto}' criado!")
    
    with sidebar_tab2:
        st.subheader("📂 Abrir Projeto")
        if st.session_state.projetos:
            projeto_selecionado = st.selectbox("Selecione um projeto", list(st.session_state.projetos.keys()), key="sb_select_proj")
            if st.button("📂 Abrir", use_container_width=True, key="btn_abrir_proj"):
                st.session_state.projeto_atual = projeto_selecionado
        else:
            st.info("Nenhum projeto disponível.")
    
    with sidebar_tab3:
        st.subheader("🔧 Ferramentas")
        col1, col2 = st.columns(2)
        with col1:
            if st.button("🗑️ Limpar", use_container_width=True, key="btn_limpar_sess"):
                st.session_state.clear()
        with col2:
            if st.button("💾 Exportar", use_container_width=True, key="btn_exp_hist"):
                if st.session_state.historico_calculos:
                    df_hist = pd.DataFrame(st.session_state.historico_calculos)
                    csv = df_hist.to_csv(index=False)
                    st.download_button("📥 CSV", csv, "historico.csv", "text/csv", key="dl_csv_hist")
    
    # Informações do projeto atual
    st.divider()
    if st.session_state.projeto_atual:
        projeto = st.session_state.projetos[st.session_state.projeto_atual]
        st.markdown(f"**📌 Ativo:** {projeto['nome']}")
        st.caption(f"👤 {projeto['cliente']}")
        st.caption(f"📍 {projeto['local']}")

# ============ INTERFACE STREAMLIT ============

st.title("⚡ Software Profissional para Projetos Elétricos")

if st.session_state.projeto_atual:
    projeto = st.session_state.projetos[st.session_state.projeto_atual]
    st.markdown(f"### 📋 {projeto['nome']}")
    st.markdown(f"Cliente: **{projeto['cliente']}** | Local: **{projeto['local']}**")
else:
    st.warning("⚠️ Selecione ou crie um projeto na barra lateral para continuar")

st.markdown("**Dimensionamento conforme NBR 5410 / NBR 5356 / IEC 60909 / NBR 5419**")
st.divider()

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "📦 Condutores", 
    "🔋 Transformadores", 
    "⚙️ Disjuntores",
    "⚡ Curto-Circuito",
    "⚖️ Balanceamento de Fases",
    "📐 Esquema Unifilar",
    "⚡ SPDA (Descargas Atmosféricas)"
])

# ============ ABA 1: CONDUTORES (WORKFLOW PROFISSIONAL) ============
with tab1:
    st.header("⚡ Dimensionamento de Condutores (NBR 5410)")
    st.markdown("**Critério:** Queda de tensão máxima e ampacidade | Tabelas 33, 36, 42, 46")
    
    if not st.session_state.projeto_atual:
        st.warning("⚠️ Crie ou selecione um projeto para usar este módulo")
    else:
        # SEÇÃO 1: IDENTIFICAÇÃO DO CIRCUITO
        st.subheader("📍 1. Identificação do Circuito")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            nome_circuito = st.text_input("Nome do Circuito", value="Circuito 1", key="cond_1_nome")
        with col2:
            num_circuito = st.number_input("Nº do Circuito", min_value=1, value=1, key="cond_1_numero")
        with col3:
            tipo_circuito_sel = st.selectbox("Tipo de Circuito", 
                ["Terminal", "Distribuição", "Ramal", "Alimentador", "Retorno"], key="cond_1_tipo")
        with col4:
            local_circuito = st.text_input("Local/Ambiente", value="Ambiente", key="cond_1_local")
        
        # SEÇÃO 2: PARÂMETROS ELÉTRICOS
        st.subheader("⚡ 2. Parâmetros Elétricos")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            corrente_circuito = st.number_input("Corrente do Circuito (A)", min_value=0.1, value=20.0, step=0.5, key="cond_2_corrente")
        with col2:
            tensao_nominal = st.number_input("Tensão Nominal (V)", min_value=127.0, value=380.0, step=127.0, key="cond_2_tensao")
        with col3:
            queda_tensao_max = st.number_input("Queda Tensão Máx. (%)", min_value=0.1, value=3.0, step=0.5, key="cond_2_queda")
        with col4:
            material = st.selectbox("Material", ["Cobre", "Alumínio"], key="cond_2_material")
        
        # SEÇÃO 3: CARACTERÍSTICAS DE INSTALAÇÃO
        st.subheader("🏗️ 3. Características de Instalação")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            tipo_instalacao = st.selectbox(
                "Método de Instalação (NBR 5410 Tab.33)", 
                {
                    "eletroduto_embutido": "Eletroduto Embutido (B1)",
                    "cond_visivel": "Condutor Visível (A1)",
                    "eletroduto_superficie": "Eletroduto Superfície (B2)",
                    "eletrocalha": "Eletrocalha (C)",
                    "bandeja": "Bandeja (D)",
                    "enterrado": "Enterrado (E)"
                },
                key="cond_3_instalacao"
            )
        with col2:
            comprimento_circuito = st.number_input("Comprimento (m)", min_value=0.0, value=30.0, step=5.0, key="cond_3_comprimento")
        with col3:
            num_circuitos = st.number_input("Nº de Circuitos Agrupados (Tab.42)", min_value=1, max_value=9, value=1, key="cond_3_agrup")
        with col4:
            fator_temperatura = st.slider("Fator Temperatura", 0.5, 1.0, 1.0, 0.05, key="cond_3_temp")
        
        # SEÇÃO 4: PROCESSAMENTO
        st.divider()
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            btn_calcular = st.button("🔄 Calcular Dimensionamento", use_container_width=True, key="btn_condutor_calc")
        with col2:
            btn_salvar = st.button("💾 Salvar no Projeto", use_container_width=True, key="btn_condutor_salvar")
        with col3:
            st.caption("")
        
        if btn_calcular:
            resultado_cond = dimensionar_condutor(
                corrente_circuito=corrente_circuito,
                comprimento_circuito=comprimento_circuito,
                queda_tensao_max=queda_tensao_max,
                tensao_nominal=tensao_nominal,
                tipo_instalacao=tipo_instalacao,
                material=material.lower(),
                num_circuitos=num_circuitos,
                fator_temperatura=fator_temperatura
            )
            
            st.session_state['resultado_condutor'] = {
                'circuito': nome_circuito,
                'num_circuito': num_circuito,
                'resultado': resultado_cond,
                'parametros': {
                    'corrente': corrente_circuito,
                    'comprimento': comprimento_circuito,
                    'tensao': tensao_nominal,
                    'queda_max': queda_tensao_max,
                    'material': material,
                    'tipo_instalacao': tipo_instalacao
                },
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            
            st.divider()
            
            # SEÇÃO 5: RESULTADOS
            st.subheader("📊 5. Resultado do Dimensionamento")
            
            # Status geral
            if resultado_cond['conforme']:
                st.success("✅ **DIMENSIONAMENTO CONFORME COM NBR 5410**", icon="✅")
            else:
                st.error("❌ **DIMENSIONAMENTO NÃO CONFORME**", icon="❌")
            
            # Métricas principais
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Seção Calculada", f"{resultado_cond['secao_minima_calculada']:.2f} mm²")
            with col2:
                st.metric("Seção Padronizada", f"{resultado_cond['secao_selecionada']} mm²")
            with col3:
                st.metric("Ampacidade", f"{resultado_cond['ampacidade']} A")
            with col4:
                st.metric("Queda de Tensão", f"{resultado_cond['queda_tensao_real']:.2f}%")
            
            # Detalhes técnicos
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Material", resultado_cond['material'])
            with col2:
                st.metric("Método Instalação", resultado_cond['metodo_instalacao'])
            with col3:
                margem_amp = ((resultado_cond['ampacidade'] - corrente_circuito) / resultado_cond['ampacidade'] * 100)
                st.metric("Margem Ampacidade", f"{margem_amp:.1f}%")
            
            # Alertas
            if resultado_cond['alertas']:
                st.warning("⚠️ **Alertas Técnicos**")
                for alerta in resultado_cond['alertas']:
                    st.warning(f"• {alerta}")
            
            # SEÇÃO 6: EXPORTAÇÃO
            st.subheader("📥 6. Exportar Resultados")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                excel_file = exportar_excel(resultado_cond, tipo='condutor', 
                                           corrente_circuito=corrente_circuito,
                                           comprimento_circuito=comprimento_circuito,
                                           queda_tensao_max=queda_tensao_max,
                                           tensao_nominal=tensao_nominal,
                                           material=resultado_cond['material'],
                                           tipo_instalacao=resultado_cond['metodo_instalacao'],
                                           num_circuitos=num_circuitos)
                st.download_button(
                    label="📊 Excel",
                    data=excel_file,
                    file_name=f"condutor_{nome_circuito.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                relatorio = gerar_relatorio(resultado_cond, tipo='condutor', 
                                           corrente_circuito=corrente_circuito,
                                           comprimento_circuito=comprimento_circuito,
                                           queda_tensao_max=queda_tensao_max)
                st.download_button(
                    label="📄 Relatório",
                    data=relatorio,
                    file_name=f"condutor_{nome_circuito.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col3:
                if st.button("👁️ Visualizar", key="btn_vis_rel_cond", use_container_width=True):
                    with st.expander("📋 Relatório Completo"):
                        st.text(relatorio)
            
            # Histórico
            st.session_state.historico_calculos.append({
                'tipo': 'Condutor',
                'circuito': nome_circuito,
                'secao': resultado_cond['secao_selecionada'],
                'corrente': corrente_circuito,
                'queda': f"{resultado_cond['queda_tensao_real']:.2f}%",
                'conforme': 'Sim' if resultado_cond['conforme'] else 'Não',
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            })
        
        if btn_salvar:
            if 'resultado_condutor' in st.session_state:
                projeto = st.session_state.projetos[st.session_state.projeto_atual]
                if 'condutores' not in projeto['modulos']:
                    projeto['modulos']['condutores'] = []
                projeto['modulos']['condutores'].append(st.session_state['resultado_condutor'])
                st.success(f"✅ Circuito '{nome_circuito}' salvo no projeto!")
            else:
                st.warning("⚠️ Execute o cálculo antes de salvar")


# ============ ABA 2: TRANSFORMADORES (WORKFLOW PROFISSIONAL) ============
with tab2:
    st.header("🔋 Dimensionamento de Transformadores (NBR 5356)")
    st.markdown("**Objetivo:** Selecionar transformador adequado com margem de crescimento")
    
    if not st.session_state.projeto_atual:
        st.warning("⚠️ Crie ou selecione um projeto para usar este módulo")
    else:
        # SEÇÃO 1: IDENTIFICAÇÃO DO TRANSFORMADOR
        st.subheader("📍 1. Identificação do Transformador")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            nome_trafo = st.text_input("Identificação", value="Trafo Principal", key="trafo_1_nome")
        with col2:
            local_trafo = st.text_input("Local de Instalação", value="Subestação", key="trafo_1_local")
        with col3:
            tipo_trafo = st.selectbox("Tipo", ["Abaixador", "Elevador", "Isolação"], key="trafo_1_tipo")
        with col4:
            fase_trafo = st.selectbox("Fase", ["Trifásico", "Monofásico"], key="trafo_1_fase")
        
        # SEÇÃO 2: PARÂMETROS ELÉTRICOS
        st.subheader("⚡ 2. Parâmetros Elétricos")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            tensao_primaria = st.number_input("Tensão Primária (V)", min_value=100.0, value=13800.0, step=100.0, key="trafo_2_primaria")
        with col2:
            tensao_secundaria = st.number_input("Tensão Secundária (V)", min_value=100.0, value=380.0, step=10.0, key="trafo_2_secundaria")
        with col3:
            potencia_total_kw = st.number_input("Potência Estimada (kW)", min_value=0.5, value=100.0, step=5.0, key="trafo_2_potencia")
        with col4:
            fator_demanda = st.slider("Fator de Demanda", 0.5, 1.0, 0.8, 0.05, key="trafo_2_fator_dem")
        
        # SEÇÃO 3: CRESCIMENTO E SEGURANÇA
        st.subheader("📈 3. Margem de Crescimento e Segurança")
        col1, col2, col3 = st.columns(3)
        with col1:
            margem_crescimento_pct = st.slider("Margem de Crescimento (%)", 0.0, 50.0, 20.0, 5.0, key="trafo_3_margem")
            margem_crescimento = margem_crescimento_pct / 100
        with col2:
            fator_seguranca = st.slider("Fator de Segurança", 1.0, 1.5, 1.1, 0.05, key="trafo_3_seguranca")
        with col3:
            st.caption("Valores típicos:\n• Crescimento: 15-30%\n• Segurança: 1.1-1.3")
        
        # SEÇÃO 4: PROCESSAMENTO
        st.divider()
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            btn_calcular_trafo = st.button("🔄 Calcular Dimensionamento", use_container_width=True, key="btn_trafo_calc")
        with col2:
            btn_salvar_trafo = st.button("💾 Salvar no Projeto", use_container_width=True, key="btn_trafo_salvar")
        with col3:
            st.caption("")
        
        if btn_calcular_trafo:
            resultado_trafo = dimensionar_transformador(
                potencia_total_kw=potencia_total_kw,
                tensao_primaria=tensao_primaria,
                tensao_secundaria=tensao_secundaria,
                fator_demanda=fator_demanda,
                margem_crescimento=margem_crescimento
            )
            
            st.session_state['resultado_trafo'] = {
                'transformador': nome_trafo,
                'local': local_trafo,
                'resultado': resultado_trafo,
                'parametros': {
                    'tensao_primaria': tensao_primaria,
                    'tensao_secundaria': tensao_secundaria,
                    'potencia': potencia_total_kw,
                    'fator_demanda': fator_demanda,
                    'margem_crescimento': margem_crescimento_pct
                },
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            
            st.divider()
            
            # SEÇÃO 5: RESULTADO DA SELEÇÃO
            st.subheader("📊 5. Resultado da Seleção")
            
            # Status geral
            if resultado_trafo['conforme']:
                st.success("✅ **TRANSFORMADOR SELECIONADO CONFORME COM NBR 5356**", icon="✅")
            else:
                st.error("❌ **VERIFICAR ALERTAS ABAIXO**", icon="❌")
            
            # Potências
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Potência Demanda", f"{resultado_trafo['potencia_demanda']:.2f} kW")
            with col2:
                st.metric("Potência Projeto", f"{resultado_trafo['potencia_projeto']:.2f} kW")
            with col3:
                st.metric("Trafo Selecionado", f"{resultado_trafo['kva_selecionado']} kVA")
            with col4:
                margem_cap = ((resultado_trafo['kva_selecionado'] * 0.9 - resultado_trafo['potencia_demanda']) / resultado_trafo['kva_selecionado'] * 100)
                st.metric("Margem Disponível", f"{margem_cap:.1f}%")
            
            # Correntes
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Corrente Primária", f"{resultado_trafo['corrente_primaria']:.2f} A")
            with col2:
                st.metric("Corrente Secundária", f"{resultado_trafo['corrente_secundaria']:.2f} A")
            with col3:
                relacao_transformacao = tensao_primaria / tensao_secundaria
                st.metric("Relação Transformação", f"1:{relacao_transformacao:.1f}")
            
            # Alertas
            if resultado_trafo['alertas']:
                st.warning("⚠️ **Alertas Técnicos**")
                for alerta in resultado_trafo['alertas']:
                    st.warning(f"• {alerta}")
            
            # SEÇÃO 6: ANÁLISE COMPARATIVA
            st.subheader("📊 6. Análise Comparativa de Opções")
            
            opcoes_kva = [10, 15, 20, 25, 30, 37.5, 45, 50, 75, 100, 150, 200]
            opcoes_selecionadas = [kva for kva in opcoes_kva if kva >= resultado_trafo['potencia_projeto']][:5]
            
            dados_comparacao = []
            for kva_opt in opcoes_selecionadas:
                pot_disponivel = kva_opt * 0.9
                margem = ((pot_disponivel - resultado_trafo['potencia_demanda']) / pot_disponivel * 100)
                corrente_sec = (kva_opt * 1000) / (tensao_secundaria * 1.73) if fase_trafo == "Trifásico" else (kva_opt * 1000) / tensao_secundaria
                
                dados_comparacao.append({
                    'Potência (kVA)': f"{kva_opt}",
                    'Potência Disp. (kW)': f"{pot_disponivel:.1f}",
                    'Margem (%)': f"{margem:.1f}%",
                    'Corrente Sec. (A)': f"{corrente_sec:.1f}",
                    'Recomendado': '✅ SIM' if kva_opt == resultado_trafo['kva_selecionado'] else '❌'
                })
            
            df_comparacao = pd.DataFrame(dados_comparacao)
            st.dataframe(df_comparacao, use_container_width=True, hide_index=True)
            
            # SEÇÃO 7: EXPORTAÇÃO
            st.subheader("📥 7. Exportar Resultados")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                excel_file = exportar_excel(resultado_trafo, tipo='transformador',
                                           potencia_total_kw=potencia_total_kw,
                                           fator_demanda=fator_demanda,
                                           margem_crescimento=margem_crescimento_pct,
                                           tensao_primaria=tensao_primaria,
                                           tensao_secundaria=tensao_secundaria)
                st.download_button(
                    label="📊 Excel",
                    data=excel_file,
                    file_name=f"trafo_{nome_trafo.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                relatorio_trafo = gerar_relatorio(resultado_trafo, tipo='transformador',
                                                   potencia_total_kw=potencia_total_kw,
                                                   fator_demanda=fator_demanda,
                                                   margem_crescimento=margem_crescimento_pct)
                st.download_button(
                    label="📄 Relatório",
                    data=relatorio_trafo,
                    file_name=f"trafo_{nome_trafo.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col3:
                if st.button("👁️ Visualizar", key="btn_vis_rel_trafo", use_container_width=True):
                    with st.expander("📋 Relatório Completo"):
                        st.text(relatorio_trafo)
            
            # Histórico
            st.session_state.historico_calculos.append({
                'tipo': 'Transformador',
                'identificacao': nome_trafo,
                'potencia': f"{resultado_trafo['kva_selecionado']} kVA",
                'tensao_sec': tensao_secundaria,
                'conforme': 'Sim' if resultado_trafo['conforme'] else 'Não',
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            })
        
        if btn_salvar_trafo:
            if 'resultado_trafo' in st.session_state:
                projeto = st.session_state.projetos[st.session_state.projeto_atual]
                if 'transformadores' not in projeto['modulos']:
                    projeto['modulos']['transformadores'] = []
                projeto['modulos']['transformadores'].append(st.session_state['resultado_trafo'])
                st.success(f"✅ Transformador '{nome_trafo}' salvo no projeto!")
            else:
                st.warning("⚠️ Execute o cálculo antes de salvar")


# ============ ABA 3: DISJUNTORES (WORKFLOW PROFISSIONAL) ============
with tab3:
    st.header("⚙️ Seleção de Disjuntores (NBR 5410)")
    st.markdown("**Objetivo:** Selecionar protetor adequado com coordenação de seletividade")
    
    if not st.session_state.projeto_atual:
        st.warning("⚠️ Crie ou selecione um projeto para usar este módulo")
    else:
        # SEÇÃO 1: IDENTIFICAÇÃO DO CIRCUITO
        st.subheader("📍 1. Identificação do Circuito")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            nome_circuito_disj = st.text_input("Nome do Circuito", value="Circuito Protegido", key="disj_1_nome")
        with col2:
            num_circuito_disj = st.number_input("Nº do Circuito", min_value=1, value=1, key="disj_1_numero")
        with col3:
            tipo_circuito_disj = st.selectbox("Tipo", 
                ["Geral", "Iluminação", "Tomada", "Ar Condicionado", "Motor", "Fornalha"], key="disj_1_tipo")
        with col4:
            local_circuito_disj = st.text_input("Local", value="Ambiente", key="disj_1_local")
        
        # SEÇÃO 2: PARÂMETROS DO CIRCUITO
        st.subheader("⚡ 2. Parâmetros do Circuito")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            corrente_circuito_disj = st.number_input("Corrente do Circuito (A)", min_value=0.1, value=20.0, step=0.5, key="disj_2_corrente")
        with col2:
            corrente_falta = st.number_input("Corrente de Falta Estimada (A)", min_value=corrente_circuito_disj, value=corrente_circuito_disj * 10, step=10.0, key="disj_2_falta")
        with col3:
            tensao_operacao = st.number_input("Tensão de Operação (V)", min_value=127.0, value=380.0, step=127.0, key="disj_2_tensao")
        with col4:
            tipo_corrente = st.selectbox("Tipo de Corrente", ["AC", "DC"], key="disj_2_corrente_tipo")
        
        # SEÇÃO 3: CARACTERÍSTICAS DE PROTEÇÃO
        st.subheader("🛡️ 3. Características de Proteção")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            padrao = st.selectbox("Padrão de Proteção", ["B", "C", "D"], key="disj_3_padrao")
        with col2:
            corrente_nominal_custom = st.checkbox("Especificar corrente nominal", value=False, key="disj_3_custom")
            if corrente_nominal_custom:
                corrente_nominal = st.number_input("Corrente Nominal (A)", min_value=1.0, value=20.0, step=1.0, key="disj_3_corrente_nom")
            else:
                corrente_nominal = None
        with col3:
            tenpo_desligamento = st.selectbox("Tempo de Desligamento", 
                ["Instantâneo", "Curto Retardo", "Longo Retardo"], key="disj_3_tempo")
        with col4:
            capacidade_ruptura = st.number_input("Capacidade de Ruptura (kA)", min_value=1.0, value=10.0, step=1.0, key="disj_3_ruptura")
        
        # SEÇÃO 4: COORDENAÇÃO E SELETIVIDADE
        st.subheader("🔗 4. Coordenação e Seletividade")
        col1, col2, col3 = st.columns(3)
        with col1:
            usar_selectividade = st.checkbox("Aplicar critério de seletividade", value=True, key="disj_4_select")
        with col2:
            if usar_selectividade:
                corrente_upstream = st.number_input("Corrente de Proteção Montante (A)", min_value=corrente_circuito_disj * 2, value=corrente_circuito_disj * 5, key="disj_4_upstream")
            else:
                corrente_upstream = None
        with col3:
            st.caption("Seletividade garante que apenas\no protetor mais próximo\nda falta atue")
        
        # SEÇÃO 5: PROCESSAMENTO
        st.divider()
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            btn_calcular_disj = st.button("🔄 Selecionar Disjuntor", use_container_width=True, key="btn_disj_calc")
        with col2:
            btn_salvar_disj = st.button("💾 Salvar no Projeto", use_container_width=True, key="btn_disj_salvar")
        with col3:
            st.caption("")
        
        if btn_calcular_disj:
            resultado_disj = dimensionar_disjuntor(
                corrente_circuito=corrente_circuito_disj,
                tipo_circuito=tipo_circuito_disj.lower(),
                padrao=padrao.lower()
            )
            
            st.session_state['resultado_disj'] = {
                'circuito': nome_circuito_disj,
                'num_circuito': num_circuito_disj,
                'resultado': resultado_disj,
                'parametros': {
                    'corrente': corrente_circuito_disj,
                    'corrente_falta': corrente_falta,
                    'tensao': tensao_operacao,
                    'padrao': padrao,
                    'capacidade_ruptura': capacidade_ruptura
                },
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            
            st.divider()
            
            # SEÇÃO 6: RESULTADO DA SELEÇÃO
            st.subheader("📊 6. Resultado da Seleção")
            
            # Status geral
            if resultado_disj['conforme']:
                st.success("✅ **DISJUNTOR SELECIONADO CONFORME COM NBR 5410**", icon="✅")
            else:
                st.error("❌ **VERIFICAR ALERTAS ABAIXO**", icon="❌")
            
            # Especificações do protetor
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Padrão", resultado_disj['padrao'])
            with col2:
                st.metric("Corrente Nominal", f"{resultado_disj['corrente_nominal']} A")
            with col3:
                st.metric("Tipo de Protetor", resultado_disj['tipo'].title())
            with col4:
                margem_trip = ((resultado_disj['corrente_nominal'] - corrente_circuito_disj) / resultado_disj['corrente_nominal'] * 100)
                st.metric("Margem de Trip", f"{margem_trip:.1f}%")
            
            # Capacidades
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Cap. de Ruptura", f"{capacidade_ruptura} kA")
            with col2:
                st.metric("Corrente de Falta", f"{corrente_falta:.0f} A")
            with col3:
                razao_falta = corrente_falta / corrente_circuito_disj
                st.metric("Multiplicador de Falta", f"{razao_falta:.1f}x")
            
            # Análise de seletividade
            if usar_selectividade:
                st.info(f"**Seletividade:** Corrente montante = {corrente_upstream:.0f} A | "
                       f"Diferença = {corrente_upstream - resultado_disj['corrente_nominal']:.0f} A")
            
            # Alertas
            if resultado_disj['alertas']:
                st.warning("⚠️ **Alertas Técnicos**")
                for alerta in resultado_disj['alertas']:
                    st.warning(f"• {alerta}")
            
            # SEÇÃO 7: COMPARAÇÃO DE OPÇÕES
            st.subheader("📊 7. Opções de Disjuntores Disponíveis")
            
            correntes_padroes = [6, 10, 13, 16, 20, 25, 32, 40, 50, 63, 80, 100]
            opcoes_disjuntores = [c for c in correntes_padroes if c >= corrente_circuito_disj][:6]
            
            dados_opcoes = []
            for corrente_opt in opcoes_disjuntores:
                margem = ((corrente_opt - corrente_circuito_disj) / corrente_opt * 100)
                suporta_falta = "✅ SIM" if corrente_opt >= (corrente_falta / 1000 * 10) else "❌ NÃO"
                
                dados_opcoes.append({
                    'Corrente (A)': f"{corrente_opt}",
                    'Margem (%)': f"{margem:.1f}%",
                    'Suporta Falta': suporta_falta,
                    'Padrão': padrao,
                    'Recomendado': '✅ IDEAL' if corrente_opt == resultado_disj['corrente_nominal'] else ''
                })
            
            df_opcoes = pd.DataFrame(dados_opcoes)
            st.dataframe(df_opcoes, use_container_width=True, hide_index=True)
            
            # SEÇÃO 8: EXPORTAÇÃO
            st.subheader("📥 8. Exportar Resultados")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                excel_file_disj = exportar_excel(resultado_disj, tipo='disjuntor',
                                               corrente_circuito=corrente_circuito_disj,
                                               tipo_circuito=tipo_circuito_disj)
                st.download_button(
                    label="📊 Excel",
                    data=excel_file_disj,
                    file_name=f"disjuntor_{nome_circuito_disj.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                relatorio_disj = gerar_relatorio(resultado_disj, tipo='disjuntor',
                                                corrente_circuito=corrente_circuito_disj,
                                                tipo_circuito=tipo_circuito_disj)
                st.download_button(
                    label="📄 Relatório",
                    data=relatorio_disj,
                    file_name=f"disjuntor_{nome_circuito_disj.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col3:
                if st.button("👁️ Visualizar", key="btn_vis_rel_disj", use_container_width=True):
                    with st.expander("📋 Relatório Completo"):
                        st.text(relatorio_disj)
            
            # Histórico
            st.session_state.historico_calculos.append({
                'tipo': 'Disjuntor',
                'circuito': nome_circuito_disj,
                'corrente_nominal': f"{resultado_disj['corrente_nominal']} A",
                'padrao': padrao,
                'conforme': 'Sim' if resultado_disj['conforme'] else 'Não',
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            })
        
        if btn_salvar_disj:
            if 'resultado_disj' in st.session_state:
                projeto = st.session_state.projetos[st.session_state.projeto_atual]
                if 'disjuntores' not in projeto['modulos']:
                    projeto['modulos']['disjuntores'] = []
                projeto['modulos']['disjuntores'].append(st.session_state['resultado_disj'])
                st.success(f"✅ Disjuntor '{nome_circuito_disj}' salvo no projeto!")
            else:
                st.warning("⚠️ Execute o cálculo antes de salvar")


# ============ ABA 4: CURTO-CIRCUITO (WORKFLOW PROFISSIONAL) ============
with tab4:
    st.header("⚡ Análise de Corrente de Curto-Circuito (IEC 60909)")
    st.markdown("**Objetivo:** Calcular correntes de falta para dimensionar proteção e equipamentos")
    
    if not st.session_state.projeto_atual:
        st.warning("⚠️ Crie ou selecione um projeto para usar este módulo")
    else:
        # SEÇÃO 1: DADOS DO TRANSFORMADOR
        st.subheader("🔋 1. Dados da Fonte (Transformador Secundário)")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            kva_transformador = st.number_input("Potência do Trafo (kVA)", min_value=1.0, value=300.0, step=10.0, key="cc_1_kva")
        with col2:
            tensao_secundaria = st.number_input("Tensão Secundária (V)", min_value=100.0, value=380.0, step=10.0, key="cc_1_tensao_sec")
        with col3:
            uk_percent = st.number_input("Impedância Uk (%)", min_value=0.1, value=5.0, step=0.5, key="cc_1_uk")
        with col4:
            tipo_transformador = st.selectbox("Tipo Trafo", ["Dy5", "Dyn5", "Yy0", "Yz5"], key="cc_1_tipo_trafo")
        
        # SEÇÃO 2: TRAJETO DO CIRCUITO
        st.subheader("🌐 2. Trajeto do Circuito até Ponto de Falta")
        col1, col2, col3, col4 = st.columns(4)
        with col1:
            comprimento_cabo = st.number_input("Comprimento do Cabo (m)", min_value=0.0, value=0.0, step=5.0, key="cc_2_comprimento")
        with col2:
            secao_cabo = st.selectbox("Seção do Cabo (mm²)", 
                [0, 1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150, 185, 240], key="cc_2_secao")
        with col3:
            material_cabo = st.selectbox("Material", ["Cobre", "Alumínio"], key="cc_2_material")
        with col4:
            temperatura_cabo = st.slider("Temperatura do Condutor (°C)", 20, 80, 50, 5, key="cc_2_temp")
        
        # SEÇÃO 3: TIPO DE FALTA
        st.subheader("⚡ 3. Tipo de Falta Analisada")
        col1, col2, col3 = st.columns(3)
        with col1:
            tipo_curto = st.selectbox("Tipo de Curto-Circuito", 
                ["Trifásico", "Bifásico", "Monofásico Terra", "Monofásico Fase-Fase"], key="cc_3_tipo")
        with col2:
            incluir_impedancia_fonte = st.checkbox("Incluir impedância da fonte", value=True, key="cc_3_fonte")
        with col3:
            incluir_impedancia_meio = st.checkbox("Incluir impedância do meio", value=True, key="cc_3_meio")
        
        # Ajustar resistividade conforme material e temperatura
        if material_cabo == "Cobre":
            rho_base = 0.0172
        else:
            rho_base = 0.0282
        
        rho_cable = rho_base * (1 + 0.00393 * (temperatura_cabo - 20))
        x_cabo_unit = 0.00008 if material_cabo == "Cobre" else 0.00008
        
        # SEÇÃO 4: PROCESSAMENTO
        st.divider()
        col1, col2, col3 = st.columns([2, 2, 1])
        
        with col1:
            btn_calcular_cc = st.button("🔄 Calcular Corrente de Falta", use_container_width=True, key="btn_cc_calc")
        with col2:
            btn_salvar_cc = st.button("💾 Salvar Análise", use_container_width=True, key="btn_cc_salvar")
        with col3:
            st.caption("")
        
        if btn_calcular_cc:
            resultado_cc = calcular_curto_circuito(
                kva_transformador=kva_transformador,
                tensao_secundaria=tensao_secundaria,
                uk_percent=uk_percent,
                comprimento_cabo=comprimento_cabo,
                secao_cabo=secao_cabo if secao_cabo > 0 else 35,
                rho_cabo=rho_cable,
                x_cabo_unit=x_cabo_unit,
                tipo_curto=tipo_curto.lower()
            )
            
            st.session_state['resultado_cc'] = {
                'tipo_falta': tipo_curto,
                'resultado': resultado_cc,
                'parametros': {
                    'kva': kva_transformador,
                    'tensao_sec': tensao_secundaria,
                    'uk': uk_percent,
                    'comprimento': comprimento_cabo,
                    'secao': secao_cabo,
                    'material': material_cabo,
                    'temperatura': temperatura_cabo
                },
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            }
            
            st.divider()
            
            # SEÇÃO 5: RESULTADO DO CÁLCULO
            st.subheader("📊 5. Resultado do Cálculo")
            
            # Status geral
            if resultado_cc['conforme']:
                st.success("✅ **CÁLCULO CONFORME COM IEC 60909**", icon="✅")
            else:
                st.warning("⚠️ **VERIFICAR RESULTADOS**", icon="⚠️")
            
            # Correntes calculadas
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Ik\" no Secundário", f"{resultado_cc['ik_secundario']:.2f} kA")
            with col2:
                st.metric("Ik no Ponto de Falta", f"{resultado_cc['ik_ponto']:.2f} kA")
            with col3:
                st.metric("Redução por Cabo", f"{((resultado_cc['ik_secundario'] - resultado_cc['ik_ponto']) / resultado_cc['ik_secundario'] * 100):.1f}%")
            with col4:
                duracao_falta_est = 0.5  # segundos típico
                st.metric("Duração Est. Falta", f"{duracao_falta_est} s")
            
            # Impedâncias
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Impedância Trafo", f"{uk_percent}%")
            with col2:
                if comprimento_cabo > 0 and secao_cabo > 0:
                    z_cabo = np.sqrt((rho_cable * comprimento_cabo / secao_cabo) ** 2 + (x_cabo_unit * comprimento_cabo) ** 2)
                    st.metric("Impedância Cabo", f"{z_cabo:.4f} Ω")
                else:
                    st.metric("Impedância Cabo", "0 Ω")
            with col3:
                st.metric("Tipo Falta", tipo_curto)
            
            # Análise de sensibilidade
            st.subheader("📈 6. Análise de Sensibilidade")
            
            cenarios = []
            
            # Cenário 1: Redução de 20% no Uk
            resultado_uk_menor = calcular_curto_circuito(
                kva_transformador=kva_transformador,
                tensao_secundaria=tensao_secundaria,
                uk_percent=uk_percent * 0.8,
                comprimento_cabo=comprimento_cabo,
                secao_cabo=secao_cabo if secao_cabo > 0 else 35,
                rho_cabo=rho_cable,
                x_cabo_unit=x_cabo_unit,
                tipo_curto=tipo_curto.lower()
            )
            
            # Cenário 2: Com cable à origem (0 m)
            resultado_origem = calcular_curto_circuito(
                kva_transformador=kva_transformador,
                tensao_secundaria=tensao_secundaria,
                uk_percent=uk_percent,
                comprimento_cabo=0,
                secao_cabo=secao_cabo if secao_cabo > 0 else 35,
                rho_cabo=rho_cable,
                x_cabo_unit=x_cabo_unit,
                tipo_curto=tipo_curto.lower()
            )
            
            # Cenário 3: Com cable 100 m
            resultado_100m = calcular_curto_circuito(
                kva_transformador=kva_transformador,
                tensao_secundaria=tensao_secundaria,
                uk_percent=uk_percent,
                comprimento_cabo=100,
                secao_cabo=secao_cabo if secao_cabo > 0 else 35,
                rho_cabo=rho_cable,
                x_cabo_unit=x_cabo_unit,
                tipo_curto=tipo_curto.lower()
            )
            
            cenarios = [{
                'Cenário': 'Nominal (Uk reduzido 20%)',
                'Ik (kA)': f"{resultado_uk_menor['ik_ponto']:.2f}",
                'Fator': f"{resultado_uk_menor['ik_ponto'] / resultado_cc['ik_ponto']:.2f}x"
            }, {
                'Cenário': f'Pior Caso (0 m de cabo)',
                'Ik (kA)': f"{resultado_origem['ik_ponto']:.2f}",
                'Fator': f"{resultado_origem['ik_ponto'] / resultado_cc['ik_ponto']:.2f}x"
            }, {
                'Cenário': f'Melhor Caso (100 m de cabo)',
                'Ik (kA)': f"{resultado_100m['ik_ponto']:.2f}",
                'Fator': f"{resultado_100m['ik_ponto'] / resultado_cc['ik_ponto']:.2f}x"
            }]
            
            df_cenarios = pd.DataFrame(cenarios)
            st.dataframe(df_cenarios, use_container_width=True, hide_index=True)
            
            # Alertas
            if resultado_cc['alertas']:
                st.warning("⚠️ **Alertas**")
                for alerta in resultado_cc['alertas']:
                    st.warning(f"• {alerta}")
            
            # SEÇÃO 7: EXPORTAÇÃO
            st.subheader("📥 7. Exportar Resultados")
            col1, col2, col3 = st.columns(3)
            
            with col1:
                excel_file_cc = exportar_excel(resultado_cc, tipo='curto_circuito',
                                              kva_transformador=kva_transformador,
                                              tensao_secundaria=tensao_secundaria,
                                              uk_percent=uk_percent,
                                              tipo_curto=tipo_curto)
                st.download_button(
                    label="📊 Excel",
                    data=excel_file_cc,
                    file_name=f"curto_circuito_{tipo_curto.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
            
            with col2:
                relatorio_cc = gerar_relatorio(resultado_cc, tipo='curto_circuito',
                                              kva_transformador=kva_transformador,
                                              tensao_secundaria=tensao_secundaria,
                                              uk_percent=uk_percent,
                                              tipo_curto=tipo_curto)
                st.download_button(
                    label="📄 Relatório",
                    data=relatorio_cc,
                    file_name=f"curto_circuito_{tipo_curto.replace(' ', '_')}_{datetime.now().strftime('%d%m%Y')}.txt",
                    mime="text/plain",
                    use_container_width=True
                )
            
            with col3:
                if st.button("👁️ Visualizar", key="btn_vis_rel_cc", use_container_width=True):
                    with st.expander("📋 Relatório Completo"):
                        st.text(relatorio_cc)
            
            # Histórico
            st.session_state.historico_calculos.append({
                'tipo': 'Curto-circuito',
                'tipo_falta': tipo_curto,
                'ik_ponto': f"{resultado_cc['ik_ponto']:.2f} kA",
                'comprimento': comprimento_cabo,
                'conforme': 'Sim' if resultado_cc['conforme'] else 'Não',
                'timestamp': datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            })
        
        if btn_salvar_cc:
            if 'resultado_cc' in st.session_state:
                projeto = st.session_state.projetos[st.session_state.projeto_atual]
                if 'curto_circuito' not in projeto['modulos']:
                    projeto['modulos']['curto_circuito'] = []
                projeto['modulos']['curto_circuito'].append(st.session_state['resultado_cc'])
                st.success(f"✅ Análise de falta '{tipo_curto}' salva no projeto!")
            else:
                st.warning("⚠️ Execute o cálculo antes de salvar")


# ============ ABA 5: BALANCEAMENTO DE FASES ============
with tab5:
    st.header("⚖️ Balanceamento de Fases")
    st.markdown("Verificação de carga equilibrada conforme NBR 5410 (máximo 3% de desbalanceamento)")
    
    st.info("ℹ️ Insira as cargas (em kW) para cada fase. O software calculará o desbalanceamento e sugerirá redistribuição se necessário.")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.subheader("📍 Fase A")
        num_cargas_a = st.number_input("Nº de cargas (Fase A)", min_value=1, max_value=20, value=3, key="fase_a_num")
        cargas_a = []
        for i in range(num_cargas_a):
            carga = st.number_input(f"Carga A{i+1} (kW)", min_value=0.0, value=2.0, step=0.5, key=f"carga_a_{i}")
            cargas_a.append(carga)
    
    with col2:
        st.subheader("📍 Fase B")
        num_cargas_b = st.number_input("Nº de cargas (Fase B)", min_value=1, max_value=20, value=3, key="fase_b_num")
        cargas_b = []
        for i in range(num_cargas_b):
            carga = st.number_input(f"Carga B{i+1} (kW)", min_value=0.0, value=2.0, step=0.5, key=f"carga_b_{i}")
            cargas_b.append(carga)
    
    with col3:
        st.subheader("📍 Fase C")
        num_cargas_c = st.number_input("Nº de cargas (Fase C)", min_value=1, max_value=20, value=3, key="fase_c_num")
        cargas_c = []
        for i in range(num_cargas_c):
            carga = st.number_input(f"Carga C{i+1} (kW)", min_value=0.0, value=2.0, step=0.5, key=f"carga_c_{i}")
            cargas_c.append(carga)
    
    tensao_balanceamento = st.number_input("Tensão Nominal (V)", min_value=127.0, value=380.0, step=127.0, key="balanc_tensao")
    
    if st.button("Calcular Balanceamento", key="btn_balanceamento"):
        resultado_balanc = balancear_fases(cargas_a, cargas_b, cargas_c, tensao_balanceamento)
        
        st.divider()
        
        # Resultados em colunas
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Carga Fase A", f"{resultado_balanc['cargas_a']:.2f} kW")
            st.metric("Corrente Fase A", f"{resultado_balanc['correntes_a']:.2f} A")
        
        with col2:
            st.metric("Carga Fase B", f"{resultado_balanc['cargas_b']:.2f} kW")
            st.metric("Corrente Fase B", f"{resultado_balanc['correntes_b']:.2f} A")
        
        with col3:
            st.metric("Carga Fase C", f"{resultado_balanc['cargas_c']:.2f} kW")
            st.metric("Corrente Fase C", f"{resultado_balanc['correntes_c']:.2f} A")
        
        st.divider()
        
        col1, col2 = st.columns(2)
        
        with col1:
            st.metric("Desbalanceamento", f"{resultado_balanc['desbalanceamento']:.2f}%")
            st.metric("Máximo Permitido", "3.0% (NBR 5410)")
        
        with col2:
            st.metric("Corrente Média", f"{resultado_balanc['corrente_media']:.2f} A")
            st.metric("Carga Média", f"{resultado_balanc['carga_media']:.2f} kW")
        
        st.divider()
        
        # Status
        if resultado_balanc['conforme']:
            st.success("✓ Fases BALANCEADAS conforme NBR 5410!")
        else:
            st.error("⚠️ Fases DESBALANCEADAS - Redistribuição necessária")
        
        # Alertas
        if resultado_balanc['alertas']:
            for alerta in resultado_balanc['alertas']:
                st.warning(f"  • {alerta}")
        
        # Sugestão de balanceamento
        st.subheader("💡 Sugestão de Balanceamento")
        
        col1, col2, col3 = st.columns(3)
        with col1:
            desvio_a = resultado_balanc['desvio_a']
            if desvio_a > 0:
                st.info(f"Fase A: Retirar {abs(desvio_a):.2f} kW")
            else:
                st.info(f"Fase A: Adicionar {abs(desvio_a):.2f} kW")
        
        with col2:
            desvio_b = resultado_balanc['desvio_b']
            if desvio_b > 0:
                st.info(f"Fase B: Retirar {abs(desvio_b):.2f} kW")
            else:
                st.info(f"Fase B: Adicionar {abs(desvio_b):.2f} kW")
        
        with col3:
            desvio_c = resultado_balanc['desvio_c']
            if desvio_c > 0:
                st.info(f"Fase C: Retirar {abs(desvio_c):.2f} kW")
            else:
                st.info(f"Fase C: Adicionar {abs(desvio_c):.2f} kW")
        
        # Gráfico de barras
        fig, ax = plt.subplots(figsize=(10, 5))
        fases = ['Fase A', 'Fase B', 'Fase C']
        cargas = [resultado_balanc['cargas_a'], resultado_balanc['cargas_b'], resultado_balanc['cargas_c']]
        cores = ['red', 'yellow', 'blue']
        
        bars = ax.bar(fases, cargas, color=cores, alpha=0.7, edgecolor='black', linewidth=2)
        ax.axhline(y=resultado_balanc['carga_media'], color='green', linestyle='--', linewidth=2, label='Carga Média')
        ax.set_ylabel('Carga (kW)', fontsize=12, fontweight='bold')
        ax.set_title('Distribuição de Cargas por Fase', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(axis='y', alpha=0.3)
        
        # Adicionar valores nas barras
        for bar, carga in zip(bars, cargas):
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{carga:.2f} kW', ha='center', va='bottom', fontweight='bold')
        
        st.pyplot(fig)
    
    # ============ SEÇÃO: DIMENSIONAMENTO E UNIFILAR ============
    st.divider()
    st.subheader("🔧 Dimensionamento de Condutor e Unifilar")
    st.markdown("Após equilibrar as fases, dimensione o condutor conforme NBR 5410 e gere o diagrama unifilar")
    
    col1, col2 = st.columns(2)
    
    with col1:
        nome_circuito_bal = st.text_input("Nome do Circuito", value="Circuito Balanceado", key="bal_nome_circuito")
        tipo_circuito_bal = st.selectbox(
            "Tipo de Circuito",
            ["Monofásico 2 Fios", "Monofásico 3 Fios", "Trifásico Sem Neutro", "Trifásico Com Neutro", "Iluminação", "Tomada"],
            key="bal_tipo_circuito"
        )
    
    with col2:
        comprimento_circuito_bal = st.number_input(
            "Comprimento do Circuito (m)",
            min_value=1.0,
            max_value=500.0,
            value=30.0,
            step=5.0,
            key="bal_comprimento_circuito"
        )
        queda_tensao_max_bal = st.number_input(
            "Queda de Tensão Máxima (%)",
            min_value=0.5,
            max_value=10.0,
            value=3.0,
            step=0.5,
            key="bal_queda_max"
        )
    
    if st.button("📊 Dimensionar Condutor (NBR 5410)", use_container_width=True, key="btn_dimensionar_condutor_bal"):
        # Usar corrente média calculada no balanceamento
        if 'resultado_balanc' in locals():
            corrente_para_dimensionar = resultado_balanc['corrente_media']
            material_condutor_bal = st.selectbox("Material do Condutor", ["Cobre", "Alumínio"], key="bal_material_condutor")
            
            # Mapa tipo circuito para NBR 5410
            tipo_map = {
                "Monofásico 2 Fios": "monofasico_2f",
                "Monofásico 3 Fios": "monofasico_3f",
                "Trifásico Sem Neutro": "trifasico_sem_neutro",
                "Trifásico Com Neutro": "trifasico_com_neutro",
                "Iluminação": "monofasico_2f",
                "Tomada": "monofasico_2f",
            }
            
            tipo_instalacao_bal = st.selectbox(
                "Tipo de Instalação (Tabela 33 NBR 5410)",
                ["Eletroduto Embutido (B1)", "Condutor Visível (A1)", "Eletroduto Superfície (B2)", 
                 "Eletrocalha (C)", "Bandeja (D)", "Enterrado (E)"],
                index=0,
                key="bal_tipo_instalacao"
            )
            
            tipo_instalacao_map = {
                "Eletroduto Embutido (B1)": "B1",
                "Condutor Visível (A1)": "A1",
                "Eletroduto Superfície (B2)": "B2",
                "Eletrocalha (C)": "C",
                "Bandeja (D)": "D",
                "Enterrado (E)": "E",
            }
            
            resultado_condutor_bal = dimensionar_condutor(
                corrente_circuito=corrente_para_dimensionar,
                comprimento_circuito=comprimento_circuito_bal,
                material=material_condutor_bal.lower(),
                queda_tensao_max=queda_tensao_max_bal,
                tensao_nominal=resultado_balanc.get('tensao_nominal', 380.0),
                tipo_instalacao=tipo_instalacao_map[tipo_instalacao_bal],
                num_circuitos=1,
                fator_temperatura=1.0
            )
            
            st.session_state['resultado_condutor_bal'] = resultado_condutor_bal
            st.session_state['resultado_balanc_temp'] = resultado_balanc
            st.session_state['nome_circuito_bal'] = nome_circuito_bal
            st.session_state['comprimento_circuito_bal'] = comprimento_circuito_bal
    
    # Exibir resultado se disponível
    if 'resultado_condutor_bal' in st.session_state:
        resultado_condutor_bal = st.session_state['resultado_condutor_bal']
        
        st.divider()
        st.subheader("📋 Resultado do Dimensionamento")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            st.metric("Seção Selecionada", f"{resultado_condutor_bal['secao_selecionada']} mm²")
            st.metric("Material", resultado_condutor_bal['material'])
        
        with col2:
            st.metric("Ampacidade", f"{resultado_condutor_bal['ampacidade']} A")
            st.metric("Corrente Ajustada", f"{resultado_condutor_bal['corrente_ajustada']:.2f} A")
        
        with col3:
            st.metric("Queda Tensão Real", f"{resultado_condutor_bal['queda_tensao_real']:.2f}%")
            st.metric("Conformidade", "✓ OK" if resultado_condutor_bal['conforme'] else "⚠️ Revisar")
        
        if resultado_condutor_bal['alertas']:
            for alerta in resultado_condutor_bal['alertas']:
                st.warning(f"  • {alerta}")
        
        # ============ GERAR UNIFILAR ============
        st.divider()
        st.subheader("📐 Gerar Esquema Unifilar")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("🖼️ Gerar PNG (Matplotlib)", use_container_width=True, key="btn_unif_png_bal"):
                fig = gerar_unifilar_matplotlib(resultado_condutor_bal, st.session_state.get('nome_circuito_bal', 'Circuito'))
                
                img_buffer = BytesIO()
                fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
                img_buffer.seek(0)
                plt.close(fig)
                
                st.success("✓ PNG gerado com sucesso!")
                st.image(img_buffer, caption="Diagrama Unifilar (PNG)")
                
                st.download_button(
                    label="📥 Download PNG",
                    data=img_buffer.getvalue(),
                    file_name=f"Unifilar_{st.session_state.get('nome_circuito_bal', 'Circuito')}.png",
                    mime="image/png",
                    use_container_width=True
                )
        
        with col2:
            if st.button("📄 Gerar PDF", use_container_width=True, key="btn_unif_pdf_bal"):
                if REPORTLAB_AVAILABLE:
                    pdf_buffer = gerar_pdf_unifilar(resultado_condutor_bal, st.session_state.get('nome_circuito_bal', 'Circuito'))
                    st.success("✓ PDF gerado com sucesso!")
                    
                    st.download_button(
                        label="📥 Download PDF",
                        data=pdf_buffer.getvalue(),
                        file_name=f"Unifilar_{st.session_state.get('nome_circuito_bal', 'Circuito')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                else:
                    st.error("❌ ReportLab não instalado. Execute: pip install reportlab")
        
        with col3:
            if st.button("🔧 Gerar DWG (CAD)", use_container_width=True, key="btn_unif_dwg_bal"):
                if EZDXF_AVAILABLE:
                    dwg_bytes = gerar_dwg_unifilar(resultado_condutor_bal, st.session_state.get('nome_circuito_bal', 'Circuito'))
                    st.success("✓ DWG gerado com sucesso!")
                    
                    st.download_button(
                        label="📥 Download DWG",
                        data=dwg_bytes,
                        file_name=f"Unifilar_{st.session_state.get('nome_circuito_bal', 'Circuito')}.dwg",
                        mime="application/octet-stream",
                        use_container_width=True
                    )
                else:
                    st.error("❌ EzDXF não instalado. Execute: pip install ezdxf")


# ============ ABA 6: ESQUEMA UNIFILAR ============
with tab6:
    st.header("📐 Geração de Esquema Unifilar")
    st.markdown("Crie diagramas unifilares em PNG, PDF e DWG")
    
    st.info("ℹ️ Use os dados de um circuito dimensionado anteriormente para gerar o esquema unifilar.")
    
    # Entrada de dados do circuito
    col1, col2 = st.columns(2)
    
    with col1:
        nome_circuito = st.text_input("Nome do Circuito", value="Circuito Principal", key="unif_nome")
        secao_selecionada = st.selectbox("Seção do Condutor (mm²)", 
                                         [1.5, 2.5, 4, 6, 10, 16, 25, 35, 50, 70, 95, 120, 150],
                                         index=5, key="unif_secao")
    
    with col2:
        material_selecionado = st.selectbox("Material", ["Cobre", "Alumínio"], key="unif_material")
        metodo_instalacao = st.selectbox("Método de Instalação", 
                                         ["A1", "B1", "B2", "C", "D", "E"], 
                                         index=1, key="unif_metodo")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        ampacidade = st.number_input("Ampacidade (A)", min_value=1.0, value=25.0, step=1.0, key="unif_ampac")
    
    with col2:
        corrente_ajustada = st.number_input("Corrente Ajustada (A)", min_value=1.0, value=20.0, step=1.0, key="unif_corrente")
    
    with col3:
        queda_tensao = st.number_input("Queda de Tensão (%)", min_value=0.0, value=2.5, step=0.1, key="unif_queda")
    
    # Criar resultado simulado
    resultado_unifilar = {
        "secao_selecionada": secao_selecionada,
        "material": material_selecionado,
        "metodo_instalacao": metodo_instalacao,
        "ampacidade": int(ampacidade),
        "corrente_ajustada": corrente_ajustada,
        "queda_tensao_real": queda_tensao
    }
    
    st.divider()
    
    col1, col2, col3 = st.columns(3)
    
    # PNG via Matplotlib
    with col1:
        if st.button("📊 Gerar PNG", key="btn_png"):
            fig = gerar_unifilar_matplotlib(resultado_unifilar, nome_circuito)
            
            # Salvar em BytesIO
            img_buffer = BytesIO()
            fig.savefig(img_buffer, format='png', dpi=300, bbox_inches='tight')
            img_buffer.seek(0)
            
            st.pyplot(fig)
            
            st.download_button(
                label="📥 Baixar PNG",
                data=img_buffer,
                file_name=f"unifilar_{nome_circuito}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.png",
                mime="image/png"
            )
            plt.close(fig)
    
    # PDF
    with col2:
        if st.button("📄 Gerar PDF", key="btn_pdf"):
            if REPORTLAB_AVAILABLE:
                pdf_buffer = gerar_pdf_unifilar(resultado_unifilar, nome_circuito)
                
                if pdf_buffer:
                    st.success("✓ PDF gerado com sucesso!")
                    st.download_button(
                        label="📥 Baixar PDF",
                        data=pdf_buffer,
                        file_name=f"unifilar_{nome_circuito}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.pdf",
                        mime="application/pdf"
                    )
            else:
                st.error("❌ ReportLab não instalado. Execute: pip install reportlab")
    
    # DWG
    with col3:
        if st.button("🎨 Gerar DWG", key="btn_dwg"):
            if EZDXF_AVAILABLE:
                dwg_buffer = gerar_dwg_unifilar(resultado_unifilar, nome_circuito)
                
                if dwg_buffer:
                    st.success("✓ DWG gerado com sucesso!")
                    st.download_button(
                        label="📥 Baixar DWG",
                        data=dwg_buffer,
                        file_name=f"unifilar_{nome_circuito}_{datetime.now().strftime('%d%m%Y_%H%M%S')}.dwg",
                        mime="application/octet-stream"
                    )
            else:
                st.error("❌ EzDXF não instalado. Execute: pip install ezdxf")
    
    st.divider()
    
    st.subheader("📋 Informações do Esquema")
    
    info_cols = st.columns(2)
    
    with info_cols[0]:
        st.write("**Dados do Condutor:**")
        st.write(f"• Seção: {resultado_unifilar['secao_selecionada']} mm²")
        st.write(f"• Material: {resultado_unifilar['material']}")
        st.write(f"• Método: {resultado_unifilar['metodo_instalacao']}")
    
    with info_cols[1]:
        st.write("**Parâmetros Elétricos:**")
        st.write(f"• Ampacidade: {resultado_unifilar['ampacidade']} A")
        st.write(f"• Corrente: {resultado_unifilar['corrente_ajustada']:.2f} A")
        st.write(f"• Queda Tensão: {resultado_unifilar['queda_tensao_real']:.2f}%")


# ============ ABA 7: SPDA (SISTEMA DE PROTEÇÃO CONTRA DESCARGAS ATMOSFÉRICAS) ============
with tab7:
    st.header("⚡ Sistema de Proteção contra Descargas Atmosféricas (SPDA)")
    st.markdown("**Dimensionamento conforme NBR 5419:2015 - Proteção de estruturas contra descargas atmosféricas**")
    
    st.markdown("""
    ### 📋 Informações sobre SPDA
    
    O SPDA é um sistema completo que inclui:
    - **Subsistema de Captação**: Capta-raios e condutores na cobertura
    - **Subsistema de Descida**: Condutores verticais
    - **Subsistema de Aterramento**: Eletrodos de terra
    - **Equipotencialização**: Conexão de estruturas metálicas
    """)
    
    # Seção de Entrada de Dados
    st.subheader("1️⃣ Dados da Estrutura")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        classe_spda = st.selectbox(
            "Classe de Proteção (NBR 5419)",
            ["I (Máxima proteção)", "II (Alta proteção)", "III (Média proteção)", "IV (Proteção básica)"],
            index=1,
            help="I: Hospitais, data centers | II: Prédios altos, indústrias | III: Residências, comércio | IV: Estruturas temporárias"
        )
        classe_map = {'I': 'I', 'II': 'II', 'III': 'III', 'IV': 'IV'}
        classe_valor = classe_map[classe_spda.split()[0]]
        
        altura_estrutura = st.number_input(
            "Altura da Estrutura (m)",
            min_value=1.0,
            max_value=500.0,
            value=15.0,
            step=1.0,
            help="Altura total da estrutura do solo à cobertura"
        )
    
    with col2:
        comprimento_estrutura = st.number_input(
            "Comprimento (m)",
            min_value=1.0,
            max_value=500.0,
            value=40.0,
            step=1.0,
            help="Dimensão maior da base da estrutura"
        )
        
        largura_estrutura = st.number_input(
            "Largura (m)",
            min_value=1.0,
            max_value=500.0,
            value=30.0,
            step=1.0,
            help="Dimensão menor da base da estrutura"
        )
    
    with col3:
        material_spda = st.selectbox(
            "Material do SPDA",
            ["cobre", "aluminio", "aco_galvanizado", "aco_inoxidavel"],
            help="Cobre: melhor condutividade | Alumínio: leve | Aço galvanizado: custo-benefício | Aço inox: ambientes corrosivos"
        )
        
        tipo_metodo = st.selectbox(
            "Método de Proteção",
            ["esfera_rolante", "malha"],
            format_func=lambda x: "Esfera Rolante" if x == "esfera_rolante" else "Malha"
        )
    
    # Botão de cálculo
    if st.button("🔧 Dimensionar SPDA", use_container_width=True):
        resultado_spda = dimensionar_spda(
            classe_spda=classe_valor,
            altura_estrutura=altura_estrutura,
            comprimento_estrutura=comprimento_estrutura,
            largura_estrutura=largura_estrutura,
            material_spda=material_spda,
            tipo_metodo=tipo_metodo
        )
        
        # Armazenar resultado na sessão
        st.session_state['resultado_spda'] = resultado_spda
    
    # Exibir resultados se disponível
    if 'resultado_spda' in st.session_state:
        resultado_spda = st.session_state['resultado_spda']
        
        st.divider()
        st.subheader("2️⃣ Resultados do Dimensionamento")
        
        # Status de conformidade
        if resultado_spda['conforme']:
            st.success("✅ Dimensionamento conforme NBR 5419")
        else:
            st.warning("⚠️ Recomendações de conformidade:")
            for alerta in resultado_spda['alertas']:
                st.warning(f"  • {alerta}")
        
        # Resultado em colunas
        col1, col2 = st.columns(2)
        
        with col1:
            st.subheader("📊 Parâmetros de Proteção")
            
            metrics = [
                ("Nível de Proteção", resultado_spda['nivel_protecao']),
                ("Classe SPDA", resultado_spda['classe_protecao']),
                ("Eficiência Mínima", resultado_spda['eficiencia_minima']),
                ("Raio Esfera Rolante", f"{resultado_spda['raio_esfera_rolante']} m"),
                ("Tamanho da Malha", resultado_spda['tamanho_malha']),
            ]
            
            for label, value in metrics:
                st.metric(label, value)
        
        with col2:
            st.subheader("🔩 Componentes do Sistema")
            
            components = {
                "Condutores de Descida": resultado_spda['num_condutores_descida'],
                "Distância entre Condutores": f"{resultado_spda['distancia_entre_condutores']} m",
                "Anéis Condutores": resultado_spda['num_aneis_condutores'],
                "Distância entre Anéis": f"{resultado_spda['distancia_entre_aneis']} m",
                "Altura Recomendada de Captação": f"{resultado_spda['altura_captacao_recomendada']} m",
            }
            
            for label, value in components.items():
                st.write(f"**{label}:** {value}")
        
        st.divider()
        st.subheader("🧭 Materiais e Condutores")
        
        col1, col2 = st.columns(2)
        
        with col1:
            material_info = materiais_spda.get(material_spda, {})
            st.write(f"**Material Selecionado:** {resultado_spda['material_spda']}")
            st.write(f"**Espessura Mínima:** {resultado_spda['espessura_minima']} mm")
            st.write(f"**Seção Mínima Condutor:** {resultado_spda['secao_minima_condutor']} mm²")
            
            if material_info:
                st.write("**Aplicação Recomendada:**", material_info.get('aplicacao', 'Geral'))
                
                with st.expander("ℹ️ Detalhes do Material"):
                    if 'vantagens' in material_info:
                        st.write("✓ Vantagens:", ", ".join(material_info['vantagens']))
                    if 'desvantagens' in material_info:
                        st.write("✗ Desvantagens:", ", ".join(material_info['desvantagens']))
                    if 'ambientes' in material_info:
                        st.write("Ambientes recomendados:", ", ".join(material_info['ambientes']))
        
        with col2:
            st.write("**Comprimentos Necessários:**")
            st.write(f"• Condutores de Descida: {resultado_spda['comprimento_condutor_descida']} m")
            st.write(f"• Malha de Proteção: {resultado_spda['comprimento_condutor_malha']} m")
            st.write(f"• Total de Condutor: {resultado_spda['comprimento_total_condutor']} m")
            st.write(f"• Massa Aproximada: {resultado_spda['massa_aproximada_material']} kg")
            st.write(f"• Resistência de Aterramento: {resultado_spda['resistencia_aterramento_recomendada']}")
        
        st.divider()
        st.subheader("⚙️ Verificação de Equipotencialização")
        
        equipotencializacao = verificar_equipotencializacao()
        
        equipot_col1, equipot_col2 = st.columns(2)
        
        with equipot_col1:
            st.metric("Tensão de Toque Limite", f"{equipotencializacao['tensao_toque_limite']} V")
            st.metric("Corrente Segura", f"{equipotencializacao['corrente_segura']} mA")
        
        with equipot_col2:
            st.metric("Impedância do Corpo", f"{equipotencializacao['impedancia_corpo']} Ω")
            st.metric("Tempo Expo Seguro", f"{equipotencializacao['tempo_exposicao_seguro']} s")
        
        with st.expander("📋 Recomendações de Equipotencialização"):
            for i, rec in enumerate(equipotencializacao['recomendacoes'], 1):
                st.write(f"{i}. {rec}")
        
        st.divider()
        st.subheader("⚡ Estimativa de Corrente de Descarga")
        
        energia_descarga = st.slider(
            "Energia Relativa da Descarga (%)",
            min_value=10,
            max_value=200,
            value=100,
            step=10,
            help="100% = corrente média de 25 kA"
        )
        
        corrente_desc = calcular_corrente_descarga(energia_descarga)
        
        desc_col1, desc_col2, desc_col3 = st.columns(3)
        
        with desc_col1:
            st.metric("Corrente Mínima Design", f"{corrente_desc['corrente_minima_design']} kA")
        with desc_col2:
            st.metric("Corrente Média Design", f"{corrente_desc['corrente_media_design']} kA")
        with desc_col3:
            st.metric("Corrente Máxima Design", f"{corrente_desc['corrente_maxima_design']} kA")
        
        st.info(f"💡 **Corrente Estimada:** {corrente_desc['corrente_estimada']} kA | Impedância típica: {corrente_desc['impedancia_arco_tipica']} Ω")
        
        st.divider()
        st.subheader("📥 Exportar Resultados")
        
        # Preparar dados para exportação
        export_data = {
            'PARÂMETROS DE PROJETO': {
                'Classe SPDA': resultado_spda['classe_protecao'],
                'Altura da Estrutura': f"{altura_estrutura} m",
                'Comprimento': f"{comprimento_estrutura} m",
                'Largura': f"{largura_estrutura} m",
                'Material': resultado_spda['material_spda'],
                'Método': 'Esfera Rolante' if tipo_metodo == 'esfera_rolante' else 'Malha',
            },
            'RESULTADOS': {
                'Nível de Proteção': resultado_spda['nivel_protecao'],
                'Eficiência Mínima': resultado_spda['eficiencia_minima'],
                'Raio Esfera Rolante': f"{resultado_spda['raio_esfera_rolante']} m",
                'Num Condutores Descida': resultado_spda['num_condutores_descida'],
                'Num Anéis': resultado_spda['num_aneis_condutores'],
                'Comprimento Total': f"{resultado_spda['comprimento_total_condutor']} m",
                'Massa Material': f"{resultado_spda['massa_aproximada_material']} kg",
                'Resistência Aterramento': resultado_spda['resistencia_aterramento_recomendada'],
                'Conformidade': 'Conforme NBR 5419' if resultado_spda['conforme'] else 'Verificar alertas',
            }
        }
        
        # Botão de download em TXT
        export_text = "DIMENSIONAMENTO SPDA - NBR 5419:2015\n"
        export_text += "=" * 50 + "\n\n"
        
        for secao, dados in export_data.items():
            export_text += f"[{secao}]\n"
            for chave, valor in dados.items():
                export_text += f"{chave}: {valor}\n"
            export_text += "\n"
        
        if resultado_spda['alertas']:
            export_text += "[ALERTAS]\n"
            for alerta in resultado_spda['alertas']:
                export_text += f"• {alerta}\n"
        
        st.download_button(
            label="📥 Download Relatório (TXT)",
            data=export_text,
            file_name="SPDA_Dimensionamento.txt",
            mime="text/plain",
            use_container_width=True
        )


# Footer
st.divider()
st.markdown("""
---
**Software de Projetos Elétricos** | Normas: NBR 5410 • NBR 5356 • IEC 60909 • NBR 5419
""")
